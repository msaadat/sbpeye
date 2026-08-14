from __future__ import annotations

import re
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="006B3C")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MAX_CELL_LENGTH = 32_767


@dataclass(frozen=True)
class ChecklistSubject:
    """The instrument a checklist belongs to, as the export needs to describe it.

    A circular and a regulation are identified by different things — one by a reference
    and a department, the other by an edition and when we captured it — so the export
    takes the identity columns rather than a model. Everything to the right of them is
    the checklist itself and is identical for both.

    `columns` are the leading spreadsheet columns, repeated on every requirement row.
    `summary_rows` are the extra Summary-sheet fields; the checklist's own counts are
    appended by the builder.
    """

    columns: list[tuple[str, Any]]
    summary_rows: list[tuple[str, Any]]
    filename_stem: str
    column_widths: list[int] = field(default_factory=lambda: [20, 38, 16, 14])

    def safe_filename(self) -> str:
        stem = re.sub(r"[^A-Za-z0-9._-]+", "_", self.filename_stem or "").strip("._")
        return f"{stem or 'checklist'}_checklist.xlsx"


def circular_subject(circular) -> ChecklistSubject:
    date = circular.date.strftime("%Y-%m-%d") if circular.date else ""
    return ChecklistSubject(
        columns=[
            ("Circular Reference", circular.reference),
            ("Circular Title", circular.title),
            ("Department", circular.department),
            ("Circular Date", date),
        ],
        summary_rows=[
            ("Circular Reference", circular.reference),
            ("Circular Title", circular.title),
            ("Department", circular.department),
            ("Circular Date", date),
            ("Circular URL", circular.url),
        ],
        filename_stem=circular.reference or getattr(circular, "id", ""),
    )


def law_subject(document, version) -> ChecklistSubject:
    """A law's identity: what it is, which edition, and since when we have held it.

    The edition matters more here than a date does for a circular. A circular is an
    immutable dated event; a regulation is replaced in place, so a checklist is only
    true of the edition it was read from — and that has to travel with the export.
    """
    from .scraper.laws import split_law_title

    display_title, version_suffix = split_law_title(document.title)
    edition = (version.version_label if version else None) or version_suffix or "current"
    captured = (
        version.first_seen_at.strftime("%Y-%m-%d")
        if version is not None and version.first_seen_at else ""
    )
    return ChecklistSubject(
        columns=[
            ("Document Title", display_title),
            ("Type", (document.doc_type or "document").title()),
            ("Edition", edition),
            ("Captured", captured),
        ],
        summary_rows=[
            ("Document Title", display_title),
            ("Type", (document.doc_type or "document").title()),
            ("Part", document.part_label),
            ("Collection", document.parent.title if document.parent else None),
            ("Edition", edition),
            ("Captured", captured),
            ("Content Hash", version.content_hash if version else None),
            ("Source URL", (version.file_url if version else None) or document.source_url),
        ],
        filename_stem=display_title,
        column_widths=[38, 14, 24, 14],
    )


def _cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (int, float, bool)):
        return value
    text = str(value)[:MAX_CELL_LENGTH]
    if text.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


def _style_table(sheet, widths: list[int]) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_checklist_workbook(subject, checklist: dict[str, Any]) -> BytesIO:
    """Render a checklist to a workbook. `subject` is a `ChecklistSubject`.

    A model is still accepted for backward compatibility with callers that pass a
    circular directly; anything without the subject's shape is adapted here.
    """
    if not isinstance(subject, ChecklistSubject):
        subject = circular_subject(subject)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Checklist"
    headers = [
        *(header for header, _ in subject.columns),
        "Source Document", "Source Type", "Reference", "Page Start", "Page End",
        "Classification", "Requirement", "Actor", "Applicability", "Deadline",
        "Evidence", "Conditions", "Source Excerpt",
    ]
    sheet.append(headers)

    identity = [value for _, value in subject.columns]
    items = [
        item for item in checklist.get("checklist_items", [])
        if isinstance(item, dict)
    ]
    if not items:
        items = [
            {
                **unit,
                "requirement": unit.get("source_text"),
            }
            for unit in checklist.get("source_units", [])
            if isinstance(unit, dict)
            and unit.get("classification") in {"required", "optional"}
        ]
    for item in items:
        if not isinstance(item, dict):
            continue
        sheet.append([_cell_value(value) for value in [
            *identity,
            item.get("doc_label"), item.get("doc_type"), item.get("ref"),
            item.get("page_start"), item.get("page_end"), item.get("classification"),
            item.get("requirement"), item.get("actor"), item.get("applicability"),
            item.get("deadline"), item.get("evidence"), item.get("conditions"),
            item.get("source_text"),
        ]])
    _style_table(sheet, [
        *subject.column_widths,
        28, 14, 38, 11, 11, 14, 60, 24, 28, 22, 38, 38, 60,
    ])

    summary = workbook.create_sheet("Summary")
    summary.append(["Field", "Value"])
    units = [unit for unit in checklist.get("source_units", []) if isinstance(unit, dict)]
    blocks = [block for block in checklist.get("analysis_blocks", []) if isinstance(block, dict)]
    summary_rows = [
        *subject.summary_rows,
        ("Checklist Status", checklist.get("status")),
        ("Generated At", checklist.get("generated_at")),
        ("Checklist Items", len(items)),
        ("Required", sum(item.get("classification") == "required" for item in items)),
        ("Optional", sum(item.get("classification") == "optional" for item in items)),
        ("Source Units", len(units)),
        ("Analysis Blocks", len(blocks)),
        ("Coverage Gaps", len(checklist.get("coverage_gaps") or [])),
    ]
    for field, value in summary_rows:
        summary.append([field, _cell_value(value)])
    _style_table(summary, [24, 80])

    gaps = workbook.create_sheet("Coverage Gaps")
    gaps.append(["Document", "Document Type", "Reference", "Page Start", "Page End", "Reason", "Error"])
    for gap in checklist.get("coverage_gaps") or []:
        if isinstance(gap, dict):
            gaps.append([_cell_value(gap.get(key)) for key in (
                "doc_label", "doc_type", "ref", "page_start", "page_end", "reason", "error"
            )])
    _style_table(gaps, [35, 18, 38, 11, 11, 24, 60])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
