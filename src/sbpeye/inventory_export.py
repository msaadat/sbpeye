"""Render an inventory search result as a reviewable workbook.

Deliberately a sibling of `sbpeye.inventory`, not a member of it: the package is
transport-neutral by design (plan section 9) and must never import openpyxl. This
module is an adapter — it reads an `InventorySearchResponse` and writes cells.

The Coverage sheet is not decoration. A spreadsheet outlives the run that produced it,
and once detached from its provenance a partial inventory reads as an exhaustive one.
Plan section 1.2 allows a completeness claim only alongside the term set, retrieval
parameters and coverage gaps, so those travel in the file.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook

from .checklist_export import _cell_value, _style_table

INVENTORY_HEADERS = [
    "Reference", "Title", "Kind", "Department", "Date", "Status",
    "Matched Via", "Matched Terms", "Semantic Score", "Verdict Reason",
    "Source Document", "Source Type", "Locator", "Page Start", "Page End",
    "Extracted Text", "Verified", "Full Passage",
]
INVENTORY_WIDTHS = [
    26, 46, 10, 14, 12, 12, 16, 34, 13, 40, 28, 13, 12, 10, 10, 60, 10, 70,
]


def _locator_label(evidence) -> str:
    """Page for PDF attachments, character offset for HTML bodies (section 8.2).

    Spelled "char 1915" rather than the CLI's "@1915": a leading `@` is a formula
    injection vector, so `_cell_value` would prefix every offset in the column with an
    apostrophe, and xlsx stores that apostrophe as a literal character.
    """
    if evidence is None:
        return ""
    if evidence.locator_kind == "page" and evidence.page_start is not None:
        return f"p.{evidence.page_start}"
    if evidence.locator_kind == "offset" and evidence.source_start is not None:
        return f"char {evidence.source_start}"
    return evidence.source_ref or ""


def _result_rows(result) -> list[list]:
    """One row per evidence item, document columns repeated.

    The flat shape is what makes the sheet pivotable, and it matches the checklist
    export's one-row-per-item convention. A result with no evidence still emits its
    row rather than vanishing.
    """
    head = [
        result.reference or "", result.title or "", result.result_kind,
        result.department or "", result.date or "", result.status or "",
        ", ".join(result.matched_via), ", ".join(result.matched_terms),
        round(result.semantic_score, 4) if result.semantic_score else "",
        result.judge_reason or "",
    ]
    if not result.evidence:
        return [head + ["", "", "", "", "", "", "", ""]]
    return [
        head + [
            evidence.source_label or "", evidence.source_kind,
            _locator_label(evidence), evidence.page_start, evidence.page_end,
            evidence.extracted_text or "",
            "yes" if evidence.extraction_verified else "no",
            evidence.passage or "",
        ]
        for evidence in result.evidence
    ]


def build_inventory_workbook(response) -> BytesIO:
    """Render an `InventorySearchResponse` into a three-sheet workbook."""
    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "Inventory"
    sheet.append(INVENTORY_HEADERS)
    for result in response.results:
        for row in _result_rows(result):
            sheet.append([_cell_value(value) for value in row])
    _style_table(sheet, INVENTORY_WIDTHS)

    _append_coverage(workbook, response)
    _append_excluded(workbook, response)
    return _to_stream(workbook)


def _append_coverage(workbook: Workbook, response) -> None:
    coverage = response.coverage
    policy = response.retrieval_policy
    sheet = workbook.create_sheet("Coverage")
    sheet.append(["Field", "Value"])

    rows = [
        ("Query", response.query),
        ("Corpus Snapshot", response.snapshot_id),
        ("Embedding Model", policy.embedding_model),
        ("Embedding Fingerprint", policy.embedding_fingerprint),
        ("Chunker Version", policy.chunker_version),
        ("Judge Model", policy.judge_model or "(adjudication skipped)"),
        ("Prompt Versions", policy.judge_prompt_version),
        ("Semantic Band", policy.semantic_band),
        ("Spans Extracted", "yes" if policy.spans_extracted else "no"),
        ("Resolved Term Set", ", ".join(policy.term_set)),
        ("Verbatim Terms", ", ".join(policy.term_set_source.get("verbatim", []))),
        ("Caller Terms", ", ".join(policy.term_set_source.get("caller", []))),
        ("Generated Terms", ", ".join(policy.term_set_source.get("generated", []))),
        ("HyDE Passage", policy.hyde_passage),
        ("Documents In Scope", coverage.logical_documents_in_scope),
        ("Candidates (Lexical)", coverage.candidates_lexical),
        ("Candidates (Semantic)", coverage.candidates_semantic),
        ("Candidates (Union)", coverage.candidates_union),
        ("Candidates Truncated", coverage.candidates_truncated),
        ("Adjudicated Included", coverage.adjudicated_included),
        ("Adjudicated Excluded", coverage.adjudicated_excluded),
        # Undetermined means "never reviewed", not "rejected". On the measured run it
        # was 37 of 97, so a reader who cannot see it is reading a different number
        # from the one the search actually produced.
        ("Adjudicated Undetermined", coverage.adjudicated_undetermined),
        ("Results Returned", response.matched_documents),
        ("Results Truncated", response.results_truncated),
        ("Vector Sources Indexed", coverage.source_units_indexed),
        ("Vector Sources Expected", coverage.source_units_expected),
        ("Lexical Documents Indexed", coverage.lexical_documents_indexed),
        ("Lexical Documents Expected", coverage.lexical_documents_expected),
        ("Stale Or Missing Index", coverage.stale_or_missing_index),
        ("Coverage Complete", "yes" if coverage.is_complete else "no"),
    ]
    for status, count in sorted(coverage.unsearchable.items()):
        rows.append((f"Unsearchable — {status}", count))
    for reason, count in sorted(coverage.excluded_by_design.items()):
        rows.append((f"Excluded by design — {reason}", count))
    for index, warning in enumerate(coverage.warnings, start=1):
        rows.append((f"Warning {index}", warning))

    for field, value in rows:
        sheet.append([field, _cell_value(value)])
    _style_table(sheet, [30, 110])


def _append_excluded(workbook: Workbook, response) -> None:
    """Rejected candidates, with reasons — never silently dropped (section 7)."""
    sheet = workbook.create_sheet("Excluded")
    sheet.append(["Reference", "Title", "Kind", "Matched Via", "Reason"])
    for excluded in response.excluded:
        sheet.append([_cell_value(value) for value in [
            excluded.reference or "", excluded.title or "", excluded.result_kind,
            ", ".join(excluded.matched_via), excluded.judge_reason or "",
        ]])
    _style_table(sheet, [26, 46, 10, 16, 60])


def _to_stream(workbook: Workbook) -> BytesIO:
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream
