import requests
import cloudscraper
from bs4 import BeautifulSoup
import re
import logging
import threading
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from sqlalchemy.orm import Session
from ..models import Attachment, Circular, CircularRelationship
from ..database import PROJECT_ROOT, collection, embedding_backend
from ..checklist import PAGE_MARKER_RE, prepare_reference_chunks
from ..search import index_circular_fts
from .clean_html import extract_sbp_text
from ..link_routing import normalize_reference, normalize_sbp_url
import uuid
from urllib.parse import unquote, urljoin, urlparse

BASE_URL = "https://www.sbp.org.pk"
ARCHIVE_BASE_URL = "https://archive.sbp.org.pk"

# The redesigned site (July 2026) serves all circulars from a single paginated
# listing at index.php?/circulars/P{offset}; the offset advances 30 at a time.
CIRCULARS_LISTING_URL_FIRST = f"{BASE_URL}/circulars/"
CIRCULARS_LISTING_URL = f"{BASE_URL}/circulars/P{{offset}}"
LISTING_PAGE_SIZE = 30
HEADERS = {
    "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
}
HTML_CACHE_DIR = PROJECT_ROOT / "cache" / "html"
ATTACHMENTS_DIR = PROJECT_ROOT / "attachments"
ATTACHMENT_EXTENSIONS = {".pdf", ".doc", ".docx", ".xls", ".xlsx"}
# Flat asset store the redesigned site consolidated most circular attachments into.
ASSET_BASE_URL = f"{BASE_URL}/assets/documents/circulars/"
_CHROMA_WRITE_LOCK = threading.Lock()


def _extract_automation_path(soup: BeautifulSoup) -> str | None:
    """Return the legacy department/year path from the hidden automationPathHolder span.

    Archived-era circular pages carry `<span id="automationPathHolder">/psd/2016/
    index.htm</span>` — a leftover of the pre-redesign URL structure. SBP's own
    front-end (`circular-inner.js`) uses it to reconstruct download links for the
    bare relative hrefs those pages emit (e.g. `href="C3-Annexure-A.pdf"`); we mirror
    that logic since it's the only source of the original per-department/year path.
    """
    holder = soup.find(id="automationPathHolder")
    if holder is None:
        return None
    text = holder.get_text().strip()
    if not text:
        return None
    text = re.sub(r"^https?://(?:www\.)?sbp\.org\.pk", "", text, flags=re.IGNORECASE)
    text = text.split("?")[0].split("#")[0]
    text = text.replace("\\", "/")
    text = re.sub(r"/[^/]*$", "/", text)
    return text.strip("/") or None


def circular_identity(reference: str | None, url: str) -> str:
    """The stable primary-key id for a circular.

    A circular's identity is its normalized reference (e.g. "BPRD CIRCULAR NO 4 OF
    2025"), so the same circular gets the same id whether it is scraped from the new
    site, from the archive, or under a different URL slug. Unreferenced circulars fall
    back to a URL-derived id.
    """
    basis = normalize_reference(reference) or url
    return str(uuid.uuid5(uuid.NAMESPACE_URL, basis))


def _get_sbp(url: str, **kwargs):
    """Fetch an SBP URL while validating every redirect target."""
    current_url = normalize_sbp_url(url)
    for _ in range(6):
        # response = requests.get(current_url, allow_redirects=False, **kwargs)
        response = cloudscraper.create_scraper().get(current_url, **kwargs)
        status_code = getattr(response, "status_code", 200)
        if status_code not in {301, 302, 303, 307, 308}:
            return response
        location = getattr(response, "headers", {}).get("location")
        response.close()
        if not location:
            raise ValueError("SBP returned a redirect without a destination.")
        current_url = normalize_sbp_url(urljoin(current_url, location))
    raise ValueError("SBP document fetch exceeded the redirect limit.")


def fetch_page(url: str) -> BeautifulSoup:
    """Fetch a URL and return a BeautifulSoup object."""
    print(f"Fetching {url}")
    # resp = requests.get(url, headers=HEADERS, timeout=50)
    resp = cloudscraper.create_scraper().get(url, headers=HEADERS, timeout=50)
    resp.raise_for_status()
    return BeautifulSoup(resp.content, "html.parser")


def fetch_page_cached(url: str, force: bool = False) -> bytes:
    """Return circular HTML from its deterministic disk cache or the network."""
    HTML_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    circular_id = str(uuid.uuid5(uuid.NAMESPACE_URL, url))
    cache_file = HTML_CACHE_DIR / f"{circular_id}.html"

    if cache_file.exists() and not force:
        return cache_file.read_bytes()

    response = _get_sbp(url, headers=HEADERS, timeout=50)
    response.raise_for_status()
    temp_file = cache_file.with_suffix(".html.part")
    temp_file.write_bytes(response.content)
    temp_file.replace(cache_file)
    return response.content


def cached_circular_html(circular) -> bytes | None:
    """Return a circular's cached detail HTML, or None if it has not been fetched."""
    cache_id = str(uuid.uuid5(uuid.NAMESPACE_URL, circular.url or ""))
    cache_file = HTML_CACHE_DIR / f"{cache_id}.html"
    return cache_file.read_bytes() if cache_file.is_file() else None


def detect_attachments(soup: BeautifulSoup, base_url: str) -> list[dict]:
    """Return unique attachment links found in circular HTML.

    SBP's redesigned site frequently links the same annexure twice under two
    different asset paths (e.g. `/assets/document/X.pdf` and
    `/assets/documents/circulars/X.pdf`) that both serve identical file content, so
    duplicates are also collapsed by filename, not just by exact URL.

    Archived-era pages instead emit a bare relative filename (e.g.
    `href="C3-Annexure-A.pdf"`, no directory component). Resolving that against the
    circular's own pretty URL produces a dead link, so it's instead resolved against
    the flat asset store, preferring the legacy department/year path from
    `automationPathHolder` when present (see `_extract_automation_path`) with the flat
    path kept as a `fallback_url` — SBP inconsistently kept files at one location or
    the other after their redesign.
    """
    found: list[dict] = []
    seen_urls: set[str] = set()
    seen_filenames: set[str] = set()
    automation_path = _extract_automation_path(soup)

    for anchor in soup.find_all("a", href=True):
        href = anchor.get("href", "").strip()
        if not href:
            continue
        if "{" in href or "}" in href:
            # Unrendered CMS template placeholder (e.g. "{Site_url}assets/...") — a
            # real URL never contains raw curly braces, so this is always a dead link.
            continue

        has_scheme = bool(re.match(r"^[a-zA-Z][a-zA-Z0-9+.-]*:", href))
        is_bare_filename = "/" not in href and not has_scheme
        fallback_url = None
        if is_bare_filename:
            try:
                flat_url = normalize_sbp_url(urljoin(ASSET_BASE_URL, href))
            except ValueError:
                continue
            if automation_path:
                try:
                    absolute_url = normalize_sbp_url(
                        urljoin(f"{ASSET_BASE_URL}{automation_path}/", href)
                    )
                except ValueError:
                    continue
                fallback_url = flat_url
            else:
                absolute_url = flat_url
        else:
            try:
                absolute_url = normalize_sbp_url(urljoin(base_url, href))
            except ValueError:
                continue
            resolved_path = urlparse(absolute_url).path
            if (
                has_scheme
                and urlparse(absolute_url).hostname == "www.sbp.org.pk"
                and not resolved_path.startswith("/assets/")
            ):
                # A pre-redesign absolute link (e.g. /dmmd/2018/C4-ANNEX-A.pdf) written
                # directly in the circular's own HTML is dead on the live site, but the
                # frozen archive mirror still serves it unchanged at the same path.
                # Links that only became absolute via urljoin against base_url (plain
                # new-site relative hrefs) don't get this fallback — they already work.
                fallback_url = normalize_sbp_url(f"{ARCHIVE_BASE_URL}{resolved_path}")

        parsed = urlparse(absolute_url)
        extension = Path(parsed.path).suffix.lower()
        if extension not in ATTACHMENT_EXTENSIONS:
            continue
        if parsed.path.lower().endswith("-u.pdf"):
            continue
        if absolute_url in seen_urls:
            continue
        seen_urls.add(absolute_url)

        filename = unquote(Path(parsed.path).name) or f"attachment{extension}"
        if filename.casefold() in seen_filenames:
            continue
        seen_filenames.add(filename.casefold())

        entry = {
            "url": absolute_url,
            "filename": filename,
            "file_type": extension.lstrip("."),
        }
        if fallback_url:
            entry["fallback_url"] = fallback_url
        found.append(entry)

    return found


def attachment_id(circular_id: str, original_url: str) -> str:
    return str(
        uuid.uuid5(uuid.NAMESPACE_URL, f"{circular_id}:{original_url}")
    )


_FILE_TYPE_MAGIC = {
    "pdf": (b"%PDF",),
    "doc": (b"\xd0\xcf\x11\xe0",),
    "xls": (b"\xd0\xcf\x11\xe0",),
    "docx": (b"PK\x03\x04",),
    "xlsx": (b"PK\x03\x04",),
}


def _content_matches_file_type(chunk: bytes, file_type: str | None) -> bool:
    """Sniff the first bytes of a download to confirm it's a real document.

    SBP serves dead attachment links as HTTP 200 with an HTML page rather than a
    404, so a successful status code alone doesn't mean the content is real.
    """
    magics = _FILE_TYPE_MAGIC.get(file_type or "")
    if not magics:
        return True
    return any(chunk.startswith(magic) for magic in magics)


def download_attachment(
    circular_id: str,
    att_info: dict,
    force: bool = False,
) -> tuple[Path | None, bool, str | None, str | None]:
    """Stream an attachment into the local cache and atomically publish it.

    Tries `att_info["url"]` first and, if its content doesn't match the expected
    file type (see `_content_matches_file_type`), falls back to
    `att_info["fallback_url"]` when present — see `detect_attachments` for why a
    circular's attachment can resolve to two different candidate locations.
    Returns the URL that actually served valid content alongside the usual
    (path, downloaded, error) tuple.
    """
    att_id = attachment_id(circular_id, att_info["url"])
    extension = f".{att_info['file_type']}" if att_info.get("file_type") else ""
    destination_dir = ATTACHMENTS_DIR / circular_id
    destination_dir.mkdir(parents=True, exist_ok=True)
    destination = destination_dir / f"{att_id}{extension}"

    if destination.exists() and not force:
        return destination, False, None, att_info["url"]

    candidates = [att_info["url"]]
    if att_info.get("fallback_url"):
        candidates.append(att_info["fallback_url"])

    temp_path = destination.with_name(f"{destination.name}.part")
    last_error: str | None = None
    for candidate_url in candidates:
        response = None
        try:
            response = _get_sbp(
                candidate_url, headers=HEADERS, timeout=60, stream=True
            )
            response.raise_for_status()
            valid = True
            with temp_path.open("wb") as output:
                for index, chunk in enumerate(
                    response.iter_content(chunk_size=1024 * 1024)
                ):
                    if not chunk:
                        continue
                    if index == 0 and not _content_matches_file_type(
                        chunk, att_info.get("file_type")
                    ):
                        valid = False
                        break
                    output.write(chunk)
            if not valid:
                temp_path.unlink(missing_ok=True)
                last_error = f"{candidate_url} did not return a valid {att_info.get('file_type')} file."
                continue
            temp_path.replace(destination)
            return destination, True, None, candidate_url
        except Exception as exc:
            temp_path.unlink(missing_ok=True)
            last_error = str(exc)
            continue
        finally:
            if response is not None:
                response.close()

    logging.warning("Failed to download attachment %s: %s", att_info["url"], last_error)
    return None, False, last_error, att_info["url"]


def _clean_pdf_pages(raw_pages: list[str]) -> list[str]:
    """Remove repeated page furniture while retaining page-local structure."""
    page_lines = [
        [re.sub(r"\s+", " ", line).strip() for line in page.splitlines() if line.strip()]
        for page in raw_pages
    ]
    edge_counts: Counter[str] = Counter()
    for lines in page_lines:
        for line in lines[:2] + lines[-2:]:
            if len(line) <= 160:
                edge_counts[line.casefold()] += 1
    repeat_threshold = max(3, (len(page_lines) + 1) // 2)
    repeated = {
        value for value, count in edge_counts.items() if count >= repeat_threshold
    }

    cleaned_pages: list[str] = []
    page_number_re = re.compile(r"^(?:page\s+)?\d+(?:\s+of\s+\d+)?$", re.IGNORECASE)
    for lines in page_lines:
        retained: list[str] = []
        last_index = len(lines) - 1
        for index, line in enumerate(lines):
            is_edge = index <= 1 or index >= last_index - 1
            if is_edge and (line.casefold() in repeated or page_number_re.fullmatch(line)):
                continue
            if retained and retained[-1].endswith("-") and line[:1].islower():
                retained[-1] = retained[-1][:-1] + line
            else:
                retained.append(line)
        cleaned_pages.append("\n".join(retained))
    return cleaned_pages


def extract_pdf_text(pdf_path: Path) -> tuple[str, str, str | None]:
    """Extract page-aware PDF text and classify image-only documents."""
    try:
        import pdfplumber

        raw_pages: list[str] = []
        with pdfplumber.open(str(pdf_path)) as pdf:
            for page in pdf.pages:
                raw_pages.append(page.extract_text() or "")

        cleaned_pages = _clean_pdf_pages(raw_pages)
        pages_text = [
            f"[[SBPEYE_PAGE:{page_number}]]\n{text.strip()}"
            for page_number, text in enumerate(cleaned_pages, start=1)
        ]

        full_text = "\n\n".join(pages_text)
        extracted_chars = sum(
            len(PAGE_MARKER_RE.sub("", value).strip()) for value in pages_text
        )
        average_chars = extracted_chars / max(len(pages_text), 1)
        status = "scanned" if average_chars < 50 else "extracted"
        return full_text, status, None
    except Exception as exc:
        logging.warning("pdfplumber failed on %s: %s", pdf_path, exc)
        return "", "error", str(exc)


def extract_xlsx_text(xlsx_path: Path) -> tuple[str, str | None]:
    """Extract sheet names and header rows from an XLSX workbook."""
    try:
        import openpyxl

        workbook = openpyxl.load_workbook(
            str(xlsx_path), read_only=True, data_only=True
        )
        parts: list[str] = []
        try:
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                parts.append(f"Sheet: {sheet_name}")
                row = next(
                    worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
                    None,
                )
                if row:
                    headers = [str(cell) for cell in row if cell is not None]
                    if headers:
                        parts.append("Headers: " + ", ".join(headers))
        finally:
            workbook.close()
        return "\n".join(parts), None
    except Exception as exc:
        logging.warning("openpyxl failed on %s: %s", xlsx_path, exc)
        return "", str(exc)


def process_attachment(
    db: Session,
    circular: Circular,
    att_info: dict,
    force_download: bool = False,
    verbose: bool = False,
) -> Attachment:
    """Download, extract, and persist one attachment idempotently."""
    att_id = att_info.get("id") or attachment_id(circular.id, att_info["url"])
    attachment = db.query(Attachment).filter(Attachment.id == att_id).first()
    if attachment is None:
        attachment = Attachment(
            id=att_id,
            circular_id=circular.id,
            filename=att_info["filename"],
            original_url=att_info["url"],
            file_type=att_info["file_type"],
            extraction_status="pending",
            is_vectorized=0,
        )
        db.add(attachment)
        db.commit()

    cached_path = (
        PROJECT_ROOT / attachment.local_path if attachment.local_path else None
    )
    complete_statuses = {"extracted", "scanned", "unsupported"}
    if (
        not force_download
        and attachment.extraction_status in complete_statuses
        and cached_path is not None
        and cached_path.exists()
    ):
        if verbose:
            print(f"    [ATT] Already processed: {attachment.filename}")
        return attachment

    local_path, downloaded, download_error, resolved_url = download_attachment(
        circular.id, att_info, force=force_download
    )
    if local_path is None:
        attachment.extraction_status = "error"
        attachment.extraction_error = download_error
        db.commit()
        return attachment

    attachment.local_path = str(local_path.relative_to(PROJECT_ROOT))
    attachment.filename = att_info["filename"]
    attachment.original_url = resolved_url
    attachment.file_type = att_info["file_type"]
    attachment.extraction_error = None

    if verbose:
        state = "Downloaded" if downloaded else "Cached"
        print(f"    [ATT] {state}: {attachment.filename}")

    if attachment.file_type == "pdf":
        text, status, error = extract_pdf_text(local_path)
        attachment.content_text = text or None
        attachment.extraction_status = status
        attachment.extraction_error = error
    elif attachment.file_type == "xlsx":
        text, error = extract_xlsx_text(local_path)
        attachment.content_text = text or None
        attachment.extraction_status = "error" if error else "extracted"
        attachment.extraction_error = error
    else:
        attachment.content_text = None
        attachment.extraction_status = "unsupported"

    attachment.is_vectorized = 0
    db.commit()
    return attachment


def fetch_attachments_for_circular(
    db: Session,
    circular: Circular,
    force_fetch: bool = False,
    force_download: bool = False,
    verbose: bool = False,
) -> list[Attachment]:
    raw_html = fetch_page_cached(circular.url, force=force_fetch)
    soup = BeautifulSoup(raw_html, "html.parser")
    detected = detect_attachments(soup, circular.url)
    if verbose:
        print(f"  [ATT] Detected {len(detected)} attachment(s)")

    processed = [
        process_attachment(
            db,
            circular,
            info,
            force_download=force_download,
            verbose=verbose,
        )
        for info in detected
    ]
    circular.attachments_scanned_at = datetime.utcnow()
    db.commit()
    return processed


def _reference_year(reference: str) -> str:
    """Best-effort year extracted from a listing reference (for date parsing)."""
    match = re.search(r"\b((?:19|20)\d{2})\b", reference or "")
    return match.group(1) if match else ""


def parse_circular_listing(soup: BeautifulSoup) -> list[dict]:
    """Parse one circulars listing page into circular descriptors.

    Each entry on the redesigned site is a ``div.publication-box-new`` holding a title
    link (the slug detail URL), a reference line, and a meta line with the date and the
    department / category / type spans.
    """
    circulars: list[dict] = []
    for box in soup.select("div.publication-box-new"):
        link = box.select_one("h4 a[href]")
        if not link:
            continue
        title = re.sub(r"\s+", " ", link.get_text(strip=True))
        if not title:
            continue
        try:
            url = normalize_sbp_url(urljoin(BASE_URL, link.get("href", "").strip()))
        except ValueError:
            continue

        ref_el = box.select_one("p.mb-3.date")
        reference = re.sub(r"\s+", " ", ref_el.get_text(" ", strip=True)) if ref_el else ""

        # The reference line and the meta line are both <p class="... date">; the meta
        # line is the one carrying the department/category/type spans.
        meta = next((p for p in box.select("p.date") if p.select_one("span.dept")), None)
        department = category = doc_type = date_text = ""
        if meta is not None:
            dept_el = meta.select_one("span.dept")
            cat_el = meta.select_one("span.cat")
            type_el = meta.select_one("span.type")
            department = dept_el.get_text(strip=True) if dept_el else ""
            category = cat_el.get_text(strip=True) if cat_el else ""
            doc_type = type_el.get_text(strip=True) if type_el else ""
            # The date is the leading text node before the first "|" separator.
            date_text = meta.get_text(" ", strip=True).split("|", 1)[0].strip()

        if not department and reference:
            department = reference.split()[0]

        circulars.append({
            "reference": reference,
            "date": date_text,
            "title": title,
            "url": url,
            "department": department,
            "category": category,
            "doc_type": doc_type,
            "year": _reference_year(reference) or _reference_year(date_text),
        })
    return circulars


def _listing_total_pages(soup: BeautifulSoup) -> int:
    """Number of listing pages, read from the pagination widget with fallbacks."""
    pager = soup.select_one(".pagination-custom[data-total-pages]")
    if pager and pager.get("data-total-pages", "").isdigit():
        return max(1, int(pager["data-total-pages"]))
    total_el = soup.select_one("#total_all_records")
    if total_el and (total_el.get("value") or "").isdigit():
        total = int(total_el["value"])
        return max(1, -(-total // LISTING_PAGE_SIZE))  # ceil division
    return 1


def _latest_circular_date(db: Session) -> datetime | None:
    """Newest circular date already stored locally, if any."""
    return db.query(Circular.date).filter(Circular.date.is_not(None)).order_by(
        Circular.date.desc()
    ).limit(1).scalar()


def _oldest_listing_date(items: list[dict]) -> datetime | None:
    """Oldest parsed listing date on one listing page."""
    dates = [
        parsed
        for item in items
        if (parsed := _parse_listing_date(item.get("date", ""), item.get("year", "")))
    ]
    return min(dates) if dates else None


def discover_circulars(
    limit: int = 0,
    max_pages: int = 0,
    verbose: bool = False,
    stop_at_date: datetime | None = None,
    full_listing: bool = False,
) -> list[dict]:
    """Crawl the unified circulars listing and return circular descriptors.

    Pages are fetched in order (P0, P30, ...) until the listing is exhausted, ``limit``
    circulars have been collected, ``max_pages`` pages have been read, or the oldest
    circular on a page reaches ``stop_at_date``.
    """
    first = fetch_page(CIRCULARS_LISTING_URL_FIRST)
    total_pages = _listing_total_pages(first)
    if max_pages > 0:
        total_pages = min(total_pages, max_pages)
    if verbose:
        print(f"Listing has {total_pages} page(s) to crawl")
        if stop_at_date and not full_listing:
            print(f"Stopping listing after reaching local latest date: {stop_at_date.date()}")

    results: list[dict] = []
    seen_urls: set[str] = set()
    for page_index in range(total_pages):
        soup = first if page_index == 0 else fetch_page(
            CIRCULARS_LISTING_URL.format(offset=page_index * LISTING_PAGE_SIZE)
        )
        page_items = parse_circular_listing(soup)
        if verbose:
            print(f"  [PAGE {page_index + 1}/{total_pages}] {len(page_items)} circular(s)")
        for item in page_items:
            if item["url"] in seen_urls:
                continue
            seen_urls.add(item["url"])
            results.append(item)
            if limit > 0 and len(results) >= limit:
                return results
        oldest_date = _oldest_listing_date(page_items)
        if (
            stop_at_date is not None
            and not full_listing
            and oldest_date is not None
            and oldest_date.date() <= stop_at_date.date()
        ):
            if verbose:
                print(
                    f"  [STOP] Page oldest date {oldest_date.date()} reached "
                    f"local latest date {stop_at_date.date()}"
                )
            break
    return results


def _matches_department(item: dict, filters: list[str]) -> bool:
    haystack = f"{item.get('department', '')} {item.get('reference', '')}".lower()
    return any(f.lower() in haystack for f in filters)


def scrape_circulars(
    db: Session,
    departments: list[str] | None = None,
    years: list[str] | None = None,
    limit: int = 0,
    skip_llm: bool = True,
    verbose: bool = False,
    force_fetch: bool = False,
    force_download: bool = False,
    include_attachments: bool = True,
    workers: int = 4,
    full_listing: bool = False,
):
    """
    Main entry point: discovers and processes circulars one by one.

    Args:
        db: SQLAlchemy session.
        departments: Optional list of department name substrings to filter.
        years: Optional list of year strings to filter (e.g., ["2025", "2024"]).
        limit: Max circulars to process (0 = unlimited).
        skip_llm: If True, skip LLM relationship extraction.
        verbose: If True, print progress details.
    """
    # With filters we must scan the whole listing to find matches, so only push the
    # limit down into the crawler for the unfiltered "latest N" case.
    filtering = bool(departments or years)
    discovered = discover_circulars(
        limit=0 if filtering else limit,
        verbose=verbose,
        stop_at_date=None if full_listing else _latest_circular_date(db),
        full_listing=full_listing,
    )

    if departments:
        discovered = [c for c in discovered if _matches_department(c, departments)]
    if years:
        discovered = [c for c in discovered if c.get("year") in years]
    if verbose and filtering:
        print(f"Filtered to {len(discovered)} circular(s)")

    pending: list[dict] = []
    skipped = 0
    for circ_info in discovered:
        if not force_fetch and not force_download:
            existing = db.query(Circular).filter(
                Circular.id == circular_identity(circ_info.get("reference"), circ_info["url"])
            ).first()
            if existing:
                skipped += 1
                if verbose:
                    print(f"Circular {circ_info['url']} already exists. Skipping")
                continue
        pending.append(circ_info)
        if limit > 0 and len(pending) >= limit:
            break

    worker_count = max(1, workers)
    print(f"Processing {len(pending)} circular(s) with {worker_count} worker(s)")

    def process_one(circ_info: dict) -> None:
        from ..database import SessionLocal

        worker_db = SessionLocal()
        try:
            process_circular(
                worker_db,
                title=circ_info["title"],
                url=circ_info["url"],
                department=circ_info["department"],
                reference=circ_info.get("reference", ""),
                listing_date=circ_info.get("date", ""),
                year=circ_info.get("year", ""),
                skip_llm=skip_llm,
                verbose=verbose,
                force_fetch=force_fetch,
                force_download=force_download,
                include_attachments=include_attachments,
            )
        finally:
            worker_db.close()

    errors = 0
    with ThreadPoolExecutor(max_workers=worker_count) as executor:
        futures = {
            executor.submit(process_one, item): (index, item)
            for index, item in enumerate(pending, start=1)
        }
        for future in as_completed(futures):
            index, item = futures[future]
            try:
                future.result()
                print(f"[{index}/{len(pending)}] {item['title']}")
            except Exception as exc:
                errors += 1
                print(f"  [ERROR] {item['url']}: {exc}")

    print(
        f"\nScrape complete. Processed: {len(pending) - errors}, "
        f"errors: {errors}, skipped (existing): {skipped}"
    )
    return {
        "discovered": len(discovered),
        "pending": len(pending),
        "processed": len(pending) - errors,
        "errors": errors,
        "skipped": skipped,
    }


def process_circular(
    db: Session,
    title: str,
    url: str,
    department: str = "Unknown",
    reference: str = "",
    listing_date: str = "",
    year: str = "",
    skip_llm: bool = True,
    verbose: bool = False,
    force_fetch: bool = False,
    force_download: bool = False,
    include_attachments: bool = True,
    old_url: str | None = None,
):
    """Download and idempotently store a circular and its attachments."""
    if verbose:
        print(f"  Fetching: {url}")

    circular_id = circular_identity(reference, url)
    raw_html = fetch_page_cached(url, force=force_fetch)
    soup = BeautifulSoup(raw_html, "html.parser")
    content_text = extract_sbp_text(raw_html)

    if not content_text:
        if verbose:
            print(f"  [SKIP] No content")
        return

    existing = db.query(Circular).filter(Circular.id == circular_id).first()
    circular_date = None
    if listing_date:
        circular_date = _parse_listing_date(listing_date, year)

    if circular_date is None:
        print(f"  [WARN] Could not parse listing date: '{listing_date}' with year '{year}' for circular {url}")
        circular_date = _extract_date(content_text)
        print(f"  [WARN] Alternate listing date: '{circular_date}'  for circular {url}")

    if circular_date is None and existing is None:
        print(f"  [WARN] Could not extract date from content for circular {url}")

    reference = re.sub(r"\s+", " ", reference)
    title = re.sub(r"\s+", " ", title)

    if existing is None:
        circular = Circular(id=circular_id, indexed_at=datetime.utcnow())
        db.add(circular)
    else:
        circular = existing
    circular.reference = reference or circular.reference
    circular.title = title
    circular.department = department
    circular.date = circular_date or circular.date
    circular.url = url
    circular.new_url = url
    if old_url:
        circular.old_url = old_url
    circular.content_text = content_text
    db.commit()

    if verbose:
        print(f"  [DB] Saved ({len(content_text)} chars, dept={department})")

    if include_attachments:
        detected = detect_attachments(soup, url)
        if verbose:
            print(f"  [ATT] Detected {len(detected)} attachment(s)")
        for info in detected:
            process_attachment(
                db,
                circular,
                info,
                force_download=force_download,
                verbose=verbose,
            )
        circular.attachments_scanned_at = datetime.utcnow()
        db.commit()

    _index_circular(circular, verbose=verbose)
    index_circular_fts(db, circular)
    return circular


def _delete_document_chunks(
    *, circular_id: str | None = None, attachment_id_value: str | None = None
) -> None:
    if attachment_id_value:
        result = collection.get(
            where={"attachment_id": attachment_id_value}, include=["metadatas"]
        )
        ids = result.get("ids", [])
    elif circular_id:
        result = collection.get(
            where={"circular_id": circular_id}, include=["metadatas"]
        )
        ids = [
            item_id
            for item_id, metadata in zip(
                result.get("ids", []), result.get("metadatas", [])
            )
            if not (metadata or {}).get("attachment_id")
        ]
    else:
        ids = []
    if ids:
        collection.delete(ids=ids)


def circular_document(circular: Circular) -> dict:
    """The text document fed to the chunker for a circular's own HTML body."""
    return {
        "doc_id": circular.id,
        "doc_type": "circular",
        "doc_label": f"{circular.department} - {circular.display_name}",
        "text": circular.content_text or "",
        "file_type": "html",
    }


def attachment_document(attachment: Attachment) -> dict:
    """The text document fed to the chunker for an extracted attachment."""
    return {
        "doc_id": attachment.id,
        "doc_type": "attachment",
        "doc_label": attachment.filename,
        "text": attachment.content_text or "",
        "file_type": attachment.file_type or "",
    }


def _chunk_page_fields(chunk: dict) -> dict:
    """Optional page bounds, omitted entirely when the chunk has none (HTML source)."""
    fields = {}
    if chunk["page_start"]:
        fields["page_start"] = chunk["page_start"]
    if chunk["page_end"]:
        fields["page_end"] = chunk["page_end"]
    return fields


def circular_chunk_metadata(circular: Circular, chunk: dict, index: int) -> dict:
    """Chroma metadata for one chunk of a circular's body text."""
    return {
        "circular_id": circular.id,
        "doc_type": "circular",
        "title": circular.title or "",
        "url": circular.url or "",
        "department": circular.department or "",
        "chunk_index": index,
        "ref": chunk["ref"],
        "unit_id": chunk["unit_id"],
        "source_start": chunk["source_start"],
        "source_end": chunk["source_end"],
        **_chunk_page_fields(chunk),
    }


def attachment_chunk_metadata(attachment: Attachment, chunk: dict, index: int) -> dict:
    """Chroma metadata for one chunk of an attachment, tagged back to its circular."""
    return {
        "circular_id": attachment.circular_id,
        "attachment_id": attachment.id,
        "doc_type": "attachment",
        "title": attachment.filename,
        "filename": attachment.filename,
        "url": attachment.original_url,
        "department": attachment.circular.department or "",
        "chunk_index": index,
        "ref": chunk["ref"],
        "unit_id": chunk["unit_id"],
        "source_start": chunk["source_start"],
        "source_end": chunk["source_end"],
        **_chunk_page_fields(chunk),
    }


def _replace_document_chunks(document: dict, *, metadata_for, delete_kwargs: dict) -> int:
    """Re-chunk, embed, and atomically swap one document's chunks in the live collection.

    `metadata_for(chunk, index)` builds the per-chunk Chroma metadata. Returns the
    number of chunks written.
    """
    reference_chunks = prepare_reference_chunks(document)
    chunks = [item["text"] for item in reference_chunks]
    doc_id = document["doc_id"]
    chunk_ids = [f"{doc_id}__chunk_{i}" for i in range(len(chunks))]
    metadatas = [metadata_for(item, i) for i, item in enumerate(reference_chunks)]
    embeddings = embedding_backend.embed_documents(chunks)
    with _CHROMA_WRITE_LOCK:
        _delete_document_chunks(**delete_kwargs)
        collection.add(
            documents=chunks,
            embeddings=embeddings,
            metadatas=metadatas,
            ids=chunk_ids,
        )
    return len(chunks)


def _index_circular(circular: Circular, verbose: bool = False) -> None:
    """Replace one circular's Chroma chunks without touching attachments."""
    try:
        count = _replace_document_chunks(
            circular_document(circular),
            metadata_for=lambda chunk, i: circular_chunk_metadata(circular, chunk, i),
            delete_kwargs={"circular_id": circular.id},
        )
        if verbose:
            print(f"  [CHROMA] Indexed ({count} chunk(s))")
    except Exception as e:
        logging.exception("ChromaDB indexing failed for %s", circular.url)
        if verbose:
            print(f"  [CHROMA] Error: {e}")


def vectorize_attachment(
    db: Session, attachment: Attachment, verbose: bool = False
) -> bool:
    """Replace the vector chunks for one extracted attachment."""
    if not attachment.content_text or not attachment.content_text.strip():
        return False

    try:
        count = _replace_document_chunks(
            attachment_document(attachment),
            metadata_for=lambda chunk, i: attachment_chunk_metadata(attachment, chunk, i),
            delete_kwargs={"attachment_id_value": attachment.id},
        )
        attachment.is_vectorized = 1
        db.commit()
        # Attachment text feeds the circular's aggregated FTS body — refresh it.
        index_circular_fts(db, attachment.circular)
        if verbose:
            print(
                f"  [CHROMA] Indexed attachment: {attachment.filename} "
                f"({count} chunks)"
            )
        return True
    except Exception:
        attachment.is_vectorized = 0
        db.commit()
        logging.exception("ChromaDB indexing failed for attachment %s", attachment.id)
        return False


def vectorize_attachments(
    db: Session, circular: Circular, verbose: bool = False
) -> int:
    indexed = 0
    for attachment in circular.attachments:
        if attachment.is_vectorized:
            continue
        if vectorize_attachment(db, attachment, verbose=verbose):
            indexed += 1
    return indexed


def reextract_circular_from_cache(
    db: Session,
    circular: Circular,
    *,
    reindex: bool = False,
    verbose: bool = False,
) -> dict[str, int]:
    """Re-extract one circular and its PDFs using local cached files only."""
    changed = 0
    errors = 0
    indexed = 0
    cache_id = str(uuid.uuid5(uuid.NAMESPACE_URL, circular.url))
    html_path = HTML_CACHE_DIR / f"{cache_id}.html"

    if html_path.is_file():
        extracted = extract_sbp_text(html_path.read_bytes())
        if extracted and extracted != (circular.content_text or ""):
            circular.content_text = extracted
            changed += 1
            _delete_document_chunks(circular_id=circular.id)
    else:
        errors += 1
        if verbose:
            print(f"  [WARN] Missing HTML cache: {html_path}")

    for attachment in circular.attachments:
        if (attachment.file_type or "").lower() != "pdf":
            continue
        local_path = PROJECT_ROOT / attachment.local_path if attachment.local_path else None
        if not local_path or not local_path.is_file():
            errors += 1
            missing_error = "Local PDF cache is missing."
            if (
                attachment.extraction_status != "error"
                or attachment.extraction_error != missing_error
                or attachment.content_text
            ):
                attachment.content_text = None
                attachment.extraction_status = "error"
                attachment.extraction_error = missing_error
                attachment.is_vectorized = 0
                changed += 1
                _delete_document_chunks(attachment_id_value=attachment.id)
            continue
        text, status, error = extract_pdf_text(local_path)
        if text != (attachment.content_text or "") or status != attachment.extraction_status:
            attachment.content_text = text or None
            attachment.extraction_status = status
            attachment.extraction_error = error
            attachment.is_vectorized = 0
            changed += 1
            _delete_document_chunks(attachment_id_value=attachment.id)

    if changed:
        circular.compliance_checklist = None
        circular.checklist_generated_at = None
    db.commit()

    if reindex:
        _index_circular(circular, verbose=verbose)
        for attachment in circular.attachments:
            if (
                (attachment.file_type or "").lower() == "pdf"
                and attachment.extraction_status == "extracted"
                and attachment.content_text
            ):
                indexed += int(vectorize_attachment(db, attachment, verbose=verbose))
        index_circular_fts(db, circular)

    return {"changed": changed, "errors": errors, "indexed": indexed}


def _extract_date(text: str) -> datetime | None:
    """Try to extract a date from circular text."""
    # Common SBP date patterns
    patterns = [
        # "January 15, 2025" or "March 3, 2024"
        r"(\b(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2},?\s+\d{4})",
        # "15th January, 2025"
        r"(\d{1,2}(?:st|nd|rd|th)?\s+(?:January|February|March|April|May|June|July|August|September|October|November|December),?\s+\d{4})",
    ]

    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            date_str = match.group(1)
            # Clean up ordinals
            date_str = re.sub(r"(\d+)(?:st|nd|rd|th)", r"\1", date_str)
            date_str = date_str.replace(",", "")
            for fmt in ["%B %d %Y", "%d %B %Y"]:
                try:
                    return datetime.strptime(date_str, fmt)
                except ValueError:
                    continue

    return None


def _parse_listing_date(date_str: str, year: str = "") -> datetime | None:
    """Parse a date string from the circular listing table."""
    date_str = date_str.strip()
    if not date_str:
        return None

    # Append the year only if no year already appears in the date string.
    if year and not re.search(r"\b(?:19|20)\d{2}\b", date_str):
        date_str = f"{date_str}, {year}"

    clean_date_str = re.sub(r'(?<=\d)(st|nd|rd|th)', '', date_str) #14th, 2nd etc
    clean_date_str = re.sub(r"\s+,\s*", ", ", clean_date_str)   # "January 15 , 2025" -> "January 15, 2025"
    clean_date_str = re.sub(r"\s+", " ", clean_date_str)        # "January  15, 2025" -> "January 15, 2025"

    formats = [
        "%b %d, %Y",
        "%d %B %Y",
        "%B %d, %Y",
        "%B %d %Y",
        "%d-%b-%Y",
        "%d-%b-%y",
        "%d.%m.%Y",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%Y-%m-%d",
        "%d %B, %Y",
        "%d %B %Y",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(clean_date_str, fmt)
        except ValueError:
            continue

    return None
