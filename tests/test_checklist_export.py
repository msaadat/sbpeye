from types import SimpleNamespace

from openpyxl import load_workbook

from sbpeye.checklist_export import build_checklist_workbook


def test_checklist_workbook_contains_requirements_summary_and_gaps():
    circular = SimpleNamespace(
        reference="BPRD/1", title="Reporting", department="BPRD",
        date=None, url="https://www.sbp.org.pk/rules.pdf",
    )
    checklist = {
        "status": "completed_with_gaps",
        "generated_at": "2026-06-20T10:00:00",
        "coverage_gaps": [{"doc_label": "scan.pdf", "doc_type": "attachment", "reason": "scanned", "error": None}],
        "analysis_blocks": [{"block_id": "block-1"}],
        "checklist_items": [{
            "doc_label": "rules.pdf", "doc_type": "attachment", "ref": "2.1",
            "page_start": 3, "page_end": 3, "classification": "required",
            "requirement": "Submit reports.", "actor": "Banks",
            "applicability": "All banks", "deadline": "Monthly",
            "evidence": "Submission receipt", "conditions": "When active",
            "source_text": "=unsafe source",
        }],
        "source_units": [{
            "doc_label": "rules.pdf", "doc_type": "attachment", "ref": "2.1",
            "page_start": 3, "page_end": 3, "source_text": "Diagnostic only",
        }],
    }

    workbook = load_workbook(build_checklist_workbook(circular, checklist))

    assert workbook.sheetnames == ["Checklist", "Summary", "Coverage Gaps"]
    assert "Review Status" not in [cell.value for cell in workbook["Checklist"][1]]
    assert workbook["Checklist"]["J2"].value == "required"
    assert workbook["Checklist"]["K2"].value == "Submit reports."
    assert workbook["Checklist"]["Q2"].value == "'=unsafe source"
    assert workbook["Checklist"].max_row == 2
    assert workbook["Summary"]["B2"].value == "BPRD/1"
    assert workbook["Coverage Gaps"]["A2"].value == "scan.pdf"
