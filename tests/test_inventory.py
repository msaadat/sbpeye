"""Tests for exhaustive inventory search (docs/INVENTORY_SEARCH_PLAN.md).

The properties pinned here are the ones the feature's completeness claim rests on:
recall is mechanical and uncapped, generated terms can only widen the search, quoted
text is verified against stored source text, and nothing is silently dropped.

Chroma, the embedding backend, and the LLM are all stubbed, so runs are offline.
"""

import json
from datetime import datetime

import numpy as np
import pytest
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from sbpeye.database import Base
from sbpeye.models import (
    Attachment,
    Circular,
    RegDocument,
    RegDocumentVersion,
    SemanticIndexSource,
)
from sbpeye.inventory import adjudicate as adjudicate_module
from sbpeye.inventory import corpus, extract, ledger, retrieval, terms
from sbpeye.inventory.adjudicate import VERDICT_INCLUDED, VERDICT_UNDETERMINED, adjudicate
from sbpeye.inventory.extract import resolve_locator, verify_span
from sbpeye.inventory.index import SNAPSHOT_CACHE, CorpusEmbeddingSnapshot, load_snapshot
from sbpeye.inventory.schemas import (
    EmbeddingFingerprintMismatch,
    InventoryFilters,
    InvalidQuery,
    InventorySearchRequest,
)
from sbpeye.inventory.service import InventorySearchService
from sbpeye.inventory_export import INVENTORY_HEADERS
from sbpeye.search import backfill_fts, backfill_laws_fts


# --------------------------------------------------------------------- doubles


class FakeCollection:
    def __init__(self):
        self.records: dict[str, dict] = {}

    def add(self, documents, embeddings, ids, metadatas):
        for doc, emb, id_, meta in zip(documents, embeddings, ids, metadatas):
            self.records[id_] = {"document": doc, "embedding": emb, "metadata": meta}

    def get(self, ids=None, where=None, limit=None, offset=None, include=None):
        items = list(self.records.items())
        if where:
            (key, value), = where.items()
            items = [(i, r) for i, r in items if r["metadata"].get(key) == value]
        if offset:
            items = items[offset:]
        if limit is not None:
            items = items[:limit]
        include = include or []
        page = {"ids": [i for i, _ in items]}
        if "metadatas" in include:
            page["metadatas"] = [r["metadata"] for _, r in items]
        if "documents" in include:
            page["documents"] = [r["document"] for _, r in items]
        if "embeddings" in include:
            page["embeddings"] = [r["embedding"] for _, r in items]
        return page

    def delete(self, ids):
        for id_ in ids:
            self.records.pop(id_, None)

    def count(self):
        return len(self.records)


class FakeBackend:
    """Deterministic 3-d embeddings: axis 0 = 'aml', axis 1 = 'audit', axis 2 = noise."""

    @staticmethod
    def _vector(text: str):
        lowered = text.lower()
        return [
            float(lowered.count("aml") + lowered.count("money laundering")),
            float(lowered.count("audit")),
            1.0,
        ]

    def embed_documents(self, documents):
        return [self._vector(d) for d in documents]

    def embed_queries(self, queries):
        return [self._vector(q) for q in queries]


class FakeConfig:
    provider = "fastembed"
    model = "test-model"


class FakeLLM:
    """Scriptable LLM. Records prompts so tests can assert what it was asked."""

    def __init__(self, terms=None, verdicts=None, span=None, fail=False):
        self._terms = terms or []
        self._verdicts = verdicts
        self._span = span
        self._fail = fail
        self.calls: list[str] = []

    model_name = "fake-judge"

    def complete_json(self, system_prompt, user_prompt, *, json_schema):
        self.calls.append(user_prompt)
        if self._fail:
            raise RuntimeError("llm down")
        properties = json_schema.get("properties", {})
        if "terms" in properties:
            return {"terms": list(self._terms)}
        if "passage" in properties:
            return {"passage": "The bank shall maintain AML controls."}
        if "verdicts" in properties:
            ids = [
                int(line.split("]")[0][1:])
                for line in user_prompt.splitlines()
                if line.startswith("[")
            ]
            if self._verdicts is None:
                return {"verdicts": [
                    {"id": i, "discusses": True, "reason": "on topic"} for i in ids
                ]}
            return {"verdicts": [
                {"id": i, "discusses": self._verdicts, "reason": "scripted"} for i in ids
            ]}
        return {"span": self._span if self._span is not None else ""}


# ----------------------------------------------------------------- fixtures


def make_session():
    engine = create_engine(
        "sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool
    )
    Base.metadata.create_all(engine)
    return sessionmaker(bind=engine, autoflush=False)()


def add_circular(db, cid, reference, title, body, department="BPRD", year=2021):
    db.add(Circular(
        id=cid, reference=reference, title=title, department=department,
        url=f"https://example.test/{cid}", content_text=body,
        date=datetime(year, 1, 1), status="active",
    ))


def add_law(db, lid, title, body, doc_type="regulation"):
    db.add(RegDocument(id=lid, title=title, doc_type=doc_type,
                       source_url=f"https://example.test/{lid}"))
    db.add(RegDocumentVersion(
        id=f"{lid}-v1", document_id=lid, content_hash=f"h-{lid}", file_type="pdf",
        content_text=body, is_current=1, is_vectorized=1,
    ))


@pytest.fixture
def corpus_db():
    db = make_session()
    add_circular(db, "c-aml", "BPRD Circular No. 07 of 2019",
                 "Anti-Money Laundering Regulations",
                 "Banks shall implement anti-money laundering and AML controls.")
    add_circular(db, "c-audit", "BPRD Circular No. 05 of 2021",
                 "Corporate Governance",
                 "The internal audit function shall report to the audit committee.")
    add_circular(db, "c-mention", "BSD Circular No. 37 of 2001",
                 "Non Performing Loans Database",
                 "Reporting of NPLs. Branches shall also observe AML instructions "
                 "when onboarding. Remaining paragraphs concern loan classification.",
                 department="BSD", year=2001)
    add_circular(db, "c-none", "DMMD Circular No. 09 of 2019",
                 "Statutory Liquidity Requirement",
                 "Banks shall maintain the prescribed liquidity ratio.",
                 department="DMMD", year=2019)
    add_law(db, "l-aml", "AML/CFT Regulations",
            "Customer due diligence and anti-money laundering obligations apply.")
    db.commit()
    backfill_fts(db, force=True)
    backfill_laws_fts(db, force=True)
    return db


@pytest.fixture
def indexed(corpus_db):
    """Index the corpus into a fake collection and reconcile the ledger."""
    from sbpeye.checklist import prepare_index_chunks

    collection = FakeCollection()
    backend = FakeBackend()
    for circular in corpus_db.query(Circular).all():
        chunks = prepare_index_chunks({
            "doc_id": circular.id, "doc_type": "circular",
            "doc_label": circular.reference, "text": circular.content_text,
            "file_type": "html",
        })
        collection.add(
            documents=[c["text"] for c in chunks],
            embeddings=backend.embed_documents([c["embed_text"] for c in chunks]),
            ids=[f"{circular.id}__chunk_{i}" for i in range(len(chunks))],
            metadatas=[{
                "circular_id": circular.id, "doc_type": "circular",
                "title": circular.title, "ref": c["ref"],
                "source_start": c["source_start"], "source_end": c["source_end"],
            } for c in chunks],
        )
    for document in corpus_db.query(RegDocument).all():
        version = document.current_version
        chunks = prepare_index_chunks({
            "doc_id": version.id, "doc_type": "law", "doc_label": document.title,
            "text": version.content_text, "file_type": "pdf",
        })
        collection.add(
            documents=[c["text"] for c in chunks],
            embeddings=backend.embed_documents([c["embed_text"] for c in chunks]),
            ids=[f"{version.id}__chunk_{i}" for i in range(len(chunks))],
            metadatas=[{
                "kind": "law", "doc_type": "law", "document_id": document.id,
                "version_id": version.id, "title": document.title, "ref": c["ref"],
                "source_start": c["source_start"], "source_end": c["source_end"],
            } for c in chunks],
        )
    ledger.reconcile(corpus_db, collection, FakeConfig(), write=True)
    return corpus_db, collection


def make_service(collection, llm=None):
    return InventorySearchService(collection, FakeBackend(), FakeConfig(), llm=llm)


# ------------------------------------------------------------------ corpus


def test_scope_excludes_circular_backed_and_manifest_laws(corpus_db):
    db = corpus_db
    db.add(RegDocument(id="l-dupe", title="Backed by circular", doc_type="regulation",
                       circular_id="c-aml"))
    db.add(RegDocument(id="l-manifest", title="Container", doc_type="law"))
    db.add(RegDocumentVersion(id="l-manifest-v1", document_id="l-manifest",
                              content_hash="h-m", file_type="manifest",
                              content_text="", is_current=1))
    db.commit()

    scope = corpus.build_scope(db)

    assert scope.excluded_by_design["law_backed_by_circular"] == 1
    assert scope.excluded_by_design["law_manifest"] == 1
    assert ("law", "l-dupe") not in scope.logical_documents()


def test_attachment_extraction_error_is_not_a_semantic_non_match(corpus_db):
    corpus_db.add(Attachment(
        id="a-bad", circular_id="c-aml", filename="scan.pdf",
        original_url="https://example.test/a", file_type="pdf",
        content_text=None, extraction_status="error", extraction_error="ocr failed",
    ))
    corpus_db.commit()

    scope = corpus.build_scope(corpus_db)
    bad = [s for s in scope.unsearchable if s.source_id == "a-bad"]

    assert bad and bad[0].unsearchable_status == corpus.STATUS_EXTRACTION_ERROR
    assert bad[0].unsearchable_detail == "ocr failed"


# ------------------------------------------------------------------ ledger


def test_ledger_marks_sources_indexed_when_chunk_counts_match(indexed):
    db, _ = indexed
    rows = {r.source_id: r for r in db.query(SemanticIndexSource).all()}

    assert rows["c-aml"].status == corpus.STATUS_INDEXED
    assert rows["c-aml"].expected_chunks == rows["c-aml"].indexed_chunks > 0


def test_ledger_detects_a_source_whose_chunks_are_missing(indexed):
    db, collection = indexed
    for chunk_id in [i for i in collection.records if i.startswith("c-audit__")]:
        collection.delete([chunk_id])

    report = ledger.reconcile(db, collection, FakeConfig(), write=True)
    row = db.query(SemanticIndexSource).filter_by(source_id="c-audit").one()

    assert row.status == corpus.STATUS_STALE
    assert not report.is_complete


def test_ledger_reports_orphan_chunks(indexed):
    db, collection = indexed
    collection.add(documents=["ghost"], embeddings=[[0.0, 0.0, 1.0]],
                   ids=["c-deleted__chunk_0"],
                   metadatas=[{"circular_id": "c-deleted", "doc_type": "circular"}])

    report = ledger.reconcile(db, collection, FakeConfig(), write=False)

    assert report.orphan_chunks == 1
    assert not report.is_complete


def test_snapshot_id_changes_when_content_changes(indexed):
    db, collection = indexed
    before = ledger.snapshot_id(db)

    db.query(Circular).filter_by(id="c-none").one().content_text = "different text"
    db.commit()
    ledger.reconcile(db, collection, FakeConfig(), write=True)

    assert ledger.snapshot_id(db) != before


# ------------------------------------------------------------------- terms


def test_generated_terms_can_only_widen_the_set():
    llm = FakeLLM(terms=["anti-money laundering", "CDD"])

    term_set, warnings = terms.build_term_set("AML", llm=llm)

    assert "AML" in term_set.all_terms
    assert "anti-money laundering" in term_set.all_terms
    assert not warnings


def test_llm_failure_leaves_the_verbatim_query_searchable():
    term_set, warnings = terms.build_term_set("AML", llm=FakeLLM(fail=True))

    assert term_set.all_terms == ["AML"]
    assert warnings and "term generation failed" in warnings[0]


def test_degenerate_llm_response_cannot_shrink_the_term_set():
    term_set, warnings = terms.build_term_set(
        "AML", alternate_queries=["money laundering"], llm=FakeLLM(terms=["", "  ", "a"])
    )

    assert term_set.all_terms == ["AML", "money laundering"]
    assert warnings and "no usable terms returned" in warnings[0]


def test_multi_word_terms_match_as_phrases():
    assert terms.fts_match_query(["call centre"]) == '"call centre"'


class _RenamingLLM:
    """Answers with the right content under a field name of its own choosing.

    Measured against nvidia/nemotron-3-ultra on OpenRouter: under the ``json_object``
    tier, which enforces no field names, the same prompt returned ``search_terms`` on
    one call and ``terms`` on the next.
    """

    model_name = "renaming"

    def __init__(self, payload):
        self._payload = payload

    def complete_json(self, system_prompt, user_prompt, *, json_schema):
        return self._payload


def test_a_renamed_terms_field_is_still_read():
    """A good answer under the wrong key is a good answer, not a failed layer."""
    term_set, warnings = terms.build_term_set(
        "call centre",
        llm=_RenamingLLM({"search_terms": ["call centres", "call centers"]}),
    )

    assert term_set.generated == ["call centres", "call centers"]
    assert not warnings


def test_an_ambiguous_response_is_not_guessed_at():
    """Two lists means no way to tell which is the answer, so fail loudly instead."""
    term_set, warnings = terms.build_term_set(
        "call centre",
        llm=_RenamingLLM({"terms": "not a list", "notes": [], "other": []}),
    )

    assert term_set.all_terms == ["call centre"]
    assert warnings and "term generation failed" in warnings[0]


def test_a_renamed_verdict_field_is_still_read():
    entries = [("A", "passage one"), ("B", "passage two")]
    llm = _RenamingLLM({"judgments": [
        {"id": 0, "discusses": True, "reason": "yes"},
        {"id": 1, "discusses": False, "reason": "no"},
    ]})

    verdicts = adjudicate_module.adjudicate(llm, "call centre", entries, max_workers=1)

    assert [v.verdict for v in verdicts] == ["included", "excluded"]


def test_verdicts_keyed_by_passage_number_are_still_read():
    """Measured shape: the judge answered with the passage numbers as top-level keys."""
    entries = [("A", "passage one"), ("B", "passage two")]
    llm = _RenamingLLM({
        "0": {"discusses": True, "reason": "yes"},
        "1": {"discusses": False, "reason": "no"},
    })

    verdicts = adjudicate_module.adjudicate(llm, "call centre", entries, max_workers=1)

    assert [v.verdict for v in verdicts] == ["included", "excluded"]
    assert [v.reason for v in verdicts] == ["yes", "no"]


def test_a_response_that_is_neither_shape_leaves_candidates_undetermined():
    """Undetermined, never excluded — an unreadable answer is not a rejection."""
    entries = [("A", "passage one")]
    llm = _RenamingLLM({"commentary": "I could not decide."})

    verdicts = adjudicate_module.adjudicate(llm, "call centre", entries, max_workers=1)

    assert verdicts[0].verdict == VERDICT_UNDETERMINED


def test_a_renamed_span_field_is_still_read():
    passage = "The bank shall staff its call centre at all times."
    llm = _RenamingLLM({"extracted_text": "shall staff its call centre"})

    results = extract.extract_spans(llm, "call centre", [passage], max_workers=1)

    assert results[0].verified
    assert results[0].text == "shall staff its call centre"


# --------------------------------------------------------------- retrieval


def test_lexical_arm_has_no_candidate_cap(corpus_db):
    for index in range(60):
        add_circular(corpus_db, f"bulk-{index}", f"REF {index}", "Bulk",
                     "This circular concerns anti-money laundering duties.")
    corpus_db.commit()
    backfill_fts(corpus_db, force=True)

    matches = lexical = retrieval.lexical_candidates(
        corpus_db, ["anti-money laundering"], include_laws=False
    )

    assert len(matches) >= 60, "lexical arm truncated its result set"
    assert all(kind == "circular" for kind, _ in lexical)


def test_lexical_arm_finds_a_passing_mention_semantics_would_rank_last(indexed):
    db, _ = indexed

    matches = retrieval.lexical_candidates(db, ["AML"], include_laws=False)

    # c-mention is mostly about NPL reporting; its AML sentence is one line in four.
    assert ("circular", "c-mention") in matches


def test_dense_band_scores_every_chunk_and_respects_its_cap(indexed):
    _, collection = indexed
    snapshot = load_snapshot(collection, "snap")
    vectors = np.asarray([[1.0, 0.0, 0.0]], dtype=np.float32)

    band = retrieval.dense_band(snapshot, vectors, band=2)

    assert len(band) == 2
    assert ("circular", "c-aml") in band


def test_union_never_truncates_a_lexical_match_before_a_semantic_one():
    lexical = {("circular", "lex"): ["aml"]}
    dense = {("circular", "sem"): (0.99, [0])}

    candidates, truncated = retrieval.union_candidates(lexical, dense, max_candidates=1)

    assert truncated == 1
    assert candidates[0].document_id == "lex"


def test_union_records_both_arms_for_a_document_found_twice():
    lexical = {("circular", "both"): ["aml"]}
    dense = {("circular", "both"): (0.9, [3, 1])}

    candidates, _ = retrieval.union_candidates(lexical, dense, max_candidates=10)

    assert candidates[0].matched_via == {"lexical", "semantic"}
    assert candidates[0].chunk_indices == [3, 1]


# --------------------------------------------------------------- extraction


def test_verified_span_must_occur_in_the_passage():
    passage = "Banks shall implement anti-money laundering controls at onboarding."

    assert verify_span("anti-money laundering controls", passage).verified
    assert verify_span("Banks must implement AML controls", passage).verified is False


def test_span_verification_tolerates_collapsed_whitespace():
    result = verify_span("shall implement AML", "Banks\n  shall implement AML now.")

    assert result.verified and result.text == "shall implement AML"


def test_locator_is_offset_for_html_bodies_and_page_for_pdfs():
    assert resolve_locator({"source_start": 10, "ref": "Chunk 1"})["locator_kind"] == "offset"
    assert resolve_locator({"page_start": 4, "ref": "Page 4"})["locator_kind"] == "page"
    assert resolve_locator({"ref": "Chunk 1"})["locator_kind"] == "chunk"


# ------------------------------------------------------------- adjudication


def test_judge_failure_marks_candidates_undetermined_not_excluded():
    verdicts = adjudicate(FakeLLM(fail=True), "AML", [("a", "text"), ("b", "text")])

    assert [v.verdict for v in verdicts] == [VERDICT_UNDETERMINED] * 2


def test_no_llm_yields_undetermined_rather_than_a_silent_pass():
    verdicts = adjudicate(None, "AML", [("a", "text")])

    assert verdicts[0].verdict == VERDICT_UNDETERMINED


def test_adjudication_preserves_entry_order_across_batches():
    entries = [(f"doc-{i}", f"passage {i}") for i in range(30)]

    verdicts = adjudicate(FakeLLM(), "AML", entries)

    assert len(verdicts) == 30
    assert all(v.verdict == VERDICT_INCLUDED for v in verdicts)


# ------------------------------------------------------------------ service


def test_search_runs_without_any_llm(indexed):
    db, collection = indexed
    request = InventorySearchRequest(
        query="anti-money laundering", generate_terms=False, use_hyde=False,
        skip_adjudication=True, extract_spans=False,
    )

    response = make_service(collection).search(request, db)

    ids = {r.document_id for r in response.results}
    assert "c-aml" in ids and "c-mention" in ids
    assert response.retrieval_policy.term_set == ["anti-money laundering"]


def test_excluded_documents_are_returned_with_reasons(indexed):
    db, collection = indexed
    request = InventorySearchRequest(
        query="anti-money laundering", generate_terms=False, use_hyde=False,
        extract_spans=False,
    )

    response = make_service(collection, llm=FakeLLM(verdicts=False)).search(request, db)

    assert response.matched_documents == 0
    assert response.excluded, "rejected candidates must stay visible"
    assert all(e.judge_reason for e in response.excluded)


def test_extraction_falls_back_to_the_passage_when_unverifiable(indexed):
    db, collection = indexed
    llm = FakeLLM(span="a paraphrase that is nowhere in the source")
    request = InventorySearchRequest(
        query="anti-money laundering", generate_terms=False, use_hyde=False,
    )

    response = make_service(collection, llm=llm).search(request, db)

    evidence = response.results[0].evidence[0]
    assert evidence.extraction_verified is False
    assert evidence.extracted_text == evidence.passage
    assert any("could not be verified" in w for w in response.coverage.warnings)


def test_a_lexical_only_hit_still_carries_evidence(indexed):
    """Acceptance criterion 5, on the path the recall backbone actually takes.

    Only the dense arm assigns chunk indices, so with the band switched off every
    candidate arrives without them. That used to emit results with an empty evidence
    list, and hand the judge the head of the document instead of the passage that
    matched — found on the first live end-to-end run.
    """
    db, collection = indexed
    request = InventorySearchRequest(
        query="anti-money laundering", generate_terms=False, use_hyde=False,
        semantic_band=0, extract_spans=False,
    )

    response = make_service(collection, llm=FakeLLM()).search(request, db)

    assert response.results, "the lexical arm alone must still return documents"
    for result in response.results:
        assert result.matched_via == ["lexical"]
        assert result.evidence, f"{result.document_id} was returned with no evidence"
        evidence = result.evidence[0]
        assert evidence.locator_kind in {"page", "offset", "chunk"}
        assert "anti-money laundering" in evidence.passage.lower()


def test_lexical_chunks_are_located_deep_inside_a_long_document():
    """The matching chunk is found wherever it sits, not just near the start.

    This is the case that made the defect matter: a long framework whose only
    call-centre paragraph is on page 40 was being judged on its cover page.
    """
    snapshot = CorpusEmbeddingSnapshot(
        snapshot_id="s",
        chunk_ids=[f"c-long__chunk_{i}" for i in range(60)],
        metadatas=[{"circular_id": "c-long"} for _ in range(60)],
        documents=[
            "the call centre shall be staffed at all times" if i == 42
            else "unrelated provisions about liquidity ratios"
            for i in range(60)
        ],
    )
    candidate = retrieval.Candidate(
        logical_kind="circular", document_id="c-long",
        matched_via={"lexical"}, matched_terms=["call centre"],
    )

    InventorySearchService._locate_lexical_chunks(snapshot, [candidate], 3)

    assert candidate.chunk_indices == [42]


def test_locating_lexical_chunks_leaves_dense_candidates_alone():
    """A candidate the dense arm already scored keeps its own, better ordering."""
    snapshot = CorpusEmbeddingSnapshot(
        snapshot_id="s",
        chunk_ids=["c-1__chunk_0", "c-1__chunk_1"],
        metadatas=[{"circular_id": "c-1"}, {"circular_id": "c-1"}],
        documents=["call centre here", "call centre there"],
    )
    candidate = retrieval.Candidate(
        logical_kind="circular", document_id="c-1",
        matched_via={"lexical", "semantic"}, matched_terms=["call centre"],
        chunk_indices=[1],
    )

    InventorySearchService._locate_lexical_chunks(snapshot, [candidate], 3)

    assert candidate.chunk_indices == [1]


# -------------------------------------------------------------- chat adapter


def _install_inventory_backends(monkeypatch, collection):
    """Point the chat tool's lazy `sbpeye.database` imports at the fixture doubles."""
    import sbpeye.database as database

    monkeypatch.setattr(database, "collection", collection)
    monkeypatch.setattr(database, "embedding_backend", FakeBackend(), raising=False)
    monkeypatch.setattr(database, "embedding_config", FakeConfig(), raising=False)


def _offline_chat_client(terms=("money laundering",)):
    """An `AIClient` whose only LLM call is stubbed, so the tool never hits a network.

    `_inventory_tool` builds its own `AIClientAdapter(self)`, and the adapter goes
    through `_complete_json`, so stubbing that one method covers every layer the tool
    can reach.
    """
    from sbpeye.ai import AIClient, AIConfig

    client = AIClient(AIConfig())
    client._complete_json = lambda *a, **kw: json.dumps({"terms": list(terms)})
    return client


def test_chat_inventory_tool_returns_unreviewed_candidates(indexed, monkeypatch):
    """The chat adapter must not present skipped adjudication as a reviewed answer.

    With `skip_adjudication` the service marks every candidate included, so the honesty
    has to come from the payload: the model is told these are unreviewed and why.
    """
    db, collection = indexed
    _install_inventory_backends(monkeypatch, collection)

    payload = json.loads(_offline_chat_client()._inventory_tool(
        {"query": "anti-money laundering", "sources": "circulars"}, db
    ))

    assert payload["reviewed"] is False
    assert "unreviewed" in payload["note"].lower()
    assert payload["results"], "the sweep should find the AML circulars"
    assert payload["documents_matched"] >= len(payload["results"])
    assert "anti-money laundering" in payload["search_terms"]


def test_chat_inventory_tool_never_calls_the_judge(indexed, monkeypatch):
    """Adjudication is 60+ s per batch; a chat turn must not wait for it."""
    db, collection = indexed
    _install_inventory_backends(monkeypatch, collection)

    client = _offline_chat_client()
    asked: list[dict] = []

    def fake_complete_json(system_prompt, user_prompt, *, json_schema, temperature=0.0):
        asked.append(json_schema.get("properties", {}))
        return '{"terms": ["money laundering"]}'

    client._complete_json = fake_complete_json
    client._inventory_tool({"query": "anti-money laundering"}, db)

    assert asked, "term generation should still run — it is the recall backbone"
    assert all("verdicts" not in props for props in asked), "the judge was called"
    assert all("span" not in props for props in asked), "extraction was called"


def test_chat_inventory_tool_cites_circulars_but_not_laws(indexed, monkeypatch):
    """Only circular and attachment tokens are resolvable by the chat UI."""
    db, collection = indexed
    _install_inventory_backends(monkeypatch, collection)

    payload = json.loads(_offline_chat_client()._inventory_tool(
        {"query": "anti-money laundering", "sources": "all", "limit": 30}, db
    ))

    cited = [r for r in payload["results"] if "citation" in r]
    assert cited, "circular results must carry a resolvable citation token"
    assert all(r["citation"].startswith("[[circular:") for r in cited)
    for result in payload["results"]:
        if "law_type" in result:
            assert "citation" not in result, "a law token would render as dead text"


def _index_attachment(db, collection, att_id, circular_id, filename, text):
    """Add an attachment and its chunks, then re-reconcile so it is not an orphan.

    The `indexed` fixture covers circular bodies and laws only, so attachment rollup
    had no coverage at all (plan section 0.4).
    """
    from sbpeye.checklist import prepare_index_chunks

    db.add(Attachment(
        id=att_id, circular_id=circular_id, filename=filename,
        original_url=f"https://example.test/{att_id}", file_type="pdf",
        content_text=text, extraction_status="success", is_vectorized=1,
    ))
    db.commit()

    backend = FakeBackend()
    chunks = prepare_index_chunks({
        "doc_id": att_id, "doc_type": "attachment", "doc_label": filename,
        "text": text, "file_type": "pdf",
    })
    collection.add(
        documents=[c["text"] for c in chunks],
        embeddings=backend.embed_documents([c["embed_text"] for c in chunks]),
        ids=[f"{att_id}__chunk_{i}" for i in range(len(chunks))],
        metadatas=[{
            "circular_id": circular_id, "doc_type": "attachment",
            "attachment_id": att_id, "filename": filename, "ref": c["ref"],
            "source_start": c["source_start"], "source_end": c["source_end"],
        } for c in chunks],
    )
    ledger.reconcile(db, collection, FakeConfig(), write=True)
    SNAPSHOT_CACHE.clear()


def test_chat_inventory_tool_cites_the_attachment_a_passage_came_from(
    indexed, monkeypatch
):
    """Showing a filename without its token is what makes the model invent one.

    Measured in a real chat: the passage text opens with "C2-AML-CFT-Regulations.pdf.
    Page 3. …", and the model produced `[[attachment:C2-AML-CFT-Regulations.pdf|…]]`
    — a citation built out of the only identifier it had been shown.
    """
    db, collection = indexed
    # Hang it off the one circular whose own body says nothing about AML, so the
    # attachment is the only evidence there is — which also exercises the rollup of
    # attachment evidence onto its parent circular (acceptance criterion 6).
    _index_attachment(
        db, collection, "att-aml", "c-none", "AML-Annex.pdf",
        "Page 1. Anti-money laundering duties of the audit function.",
    )
    _install_inventory_backends(monkeypatch, collection)

    payload = json.loads(_offline_chat_client()._inventory_tool(
        {"query": "anti-money laundering", "sources": "circulars"}, db
    ))

    rolled_up = [r for r in payload["results"] if r.get("reference", "").startswith("DMMD")]
    assert rolled_up, "a circular found only through its attachment must still appear"
    assert rolled_up[0]["attachment_citation"] == (
        "[[attachment:att-aml|AML-Annex.pdf]]"
    )
    assert "AML-Annex.pdf" in rolled_up[0]["passage"]


def test_chat_inventory_tool_returns_every_match_by_default(indexed, monkeypatch):
    """No arbitrary row cap — a fixed ceiling makes an inventory tool return a sample."""
    db, collection = indexed
    _install_inventory_backends(monkeypatch, collection)

    payload = json.loads(_offline_chat_client()._inventory_tool(
        {"query": "anti-money laundering", "sources": "circulars"}, db
    ))

    assert payload["documents_returned"] == payload["documents_matched"]
    assert payload["complete"] is True
    assert "omitted" not in payload


def test_chat_inventory_tool_says_so_in_words_when_it_cuts_the_list(
    indexed, monkeypatch
):
    """A flag alone lets the model present a partial list as an exhaustive one."""
    db, collection = indexed
    _install_inventory_backends(monkeypatch, collection)

    payload = json.loads(_offline_chat_client()._inventory_tool(
        {"query": "anti-money laundering", "sources": "circulars", "limit": 1}, db
    ))

    assert payload["documents_returned"] == 1
    assert payload["complete"] is False
    assert payload["omitted"] >= 1
    assert "INCOMPLETE" in payload["note"]


def test_chat_inventory_tool_stops_at_the_context_budget(indexed, monkeypatch):
    """When the window is the binding constraint, it binds — and is disclosed."""
    from sbpeye.ai import AIConfig

    db, collection = indexed
    _install_inventory_backends(monkeypatch, collection)

    client = _offline_chat_client()
    client.config = AIConfig(max_context_tokens=1)
    payload = json.loads(client._inventory_tool(
        {"query": "anti-money laundering", "sources": "circulars"}, db
    ))

    # The first row always goes out, however tight the budget: an empty inventory
    # would be indistinguishable from "nothing matched".
    assert payload["documents_returned"] == 1
    assert payload["complete"] is False
    assert "INCOMPLETE" in payload["note"]


def test_chat_inventory_tool_reports_errors_rather_than_raising(indexed, monkeypatch):
    db, collection = indexed
    _install_inventory_backends(monkeypatch, collection)

    payload = json.loads(_offline_chat_client()._inventory_tool({"query": "  "}, db))

    assert "error" in payload


# ------------------------------------------------------------- xlsx export


def _workbook_for(response):
    from openpyxl import load_workbook

    from sbpeye.inventory_export import build_inventory_workbook

    return load_workbook(build_inventory_workbook(response))


def _sheet_rows(sheet):
    return [[cell.value for cell in row] for row in sheet.iter_rows()]


def test_workbook_has_a_row_per_evidence_item(indexed):
    db, collection = indexed
    request = InventorySearchRequest(
        query="anti-money laundering", generate_terms=False, use_hyde=False,
        extract_spans=False,
    )
    response = make_service(collection, llm=FakeLLM()).search(request, db)

    book = _workbook_for(response)
    rows = _sheet_rows(book["Inventory"])

    assert rows[0] == INVENTORY_HEADERS
    expected = sum(max(1, len(r.evidence)) for r in response.results)
    assert len(rows) - 1 == expected
    assert any("anti-money laundering" in (r[1] or "").lower() for r in rows[1:])


def test_workbook_carries_the_provenance_a_claim_depends_on(indexed):
    """Plan 1.2: a completeness claim is only permitted alongside these fields."""
    db, collection = indexed
    request = InventorySearchRequest(
        query="anti-money laundering", generate_terms=False, use_hyde=False,
        extract_spans=False,
    )
    response = make_service(collection, llm=FakeLLM()).search(request, db)

    fields = {row[0] for row in _sheet_rows(_workbook_for(response)["Coverage"])}

    for required in (
        "Corpus Snapshot", "Resolved Term Set", "Embedding Fingerprint",
        "Chunker Version", "Adjudicated Undetermined", "Coverage Complete",
    ):
        assert required in fields, f"{required} missing from the Coverage sheet"


def test_workbook_keeps_rejected_documents_with_their_reasons(indexed):
    """Section 7: exclusions are returned, never silently dropped."""
    db, collection = indexed
    request = InventorySearchRequest(
        query="anti-money laundering", generate_terms=False, use_hyde=False,
        extract_spans=False,
    )
    response = make_service(collection, llm=FakeLLM(verdicts=False)).search(request, db)

    rows = _sheet_rows(_workbook_for(response)["Excluded"])

    assert len(rows) - 1 == len(response.excluded) > 0
    assert all(row[4] for row in rows[1:]), "every exclusion needs a reason"


def test_workbook_locators_do_not_need_formula_escaping(indexed):
    """A leading '@' would put a literal apostrophe in every offset cell."""
    db, collection = indexed
    request = InventorySearchRequest(
        query="anti-money laundering", generate_terms=False, use_hyde=False,
        extract_spans=False,
    )
    response = make_service(collection, llm=FakeLLM()).search(request, db)

    locators = [row[12] for row in _sheet_rows(_workbook_for(response)["Inventory"])[1:]]

    assert locators, "the fixture should produce located evidence"
    assert not any((loc or "").startswith("'") for loc in locators)
    assert any((loc or "").startswith("char ") for loc in locators)


def test_workbook_neutralises_text_that_would_become_a_formula():
    """Regulatory text routinely starts with '-' or '='; Excel would evaluate it."""
    from sbpeye.checklist_export import _cell_value

    assert _cell_value("=cmd|' /c calc'!A1").startswith("'")
    assert _cell_value("-3% of paid-up capital").startswith("'")
    assert _cell_value("Ordinary text") == "Ordinary text"


def test_workbook_emits_a_row_even_when_a_result_has_no_evidence():
    """A document with nothing attached must not silently vanish from the sheet."""
    from sbpeye.inventory_export import _result_rows
    from sbpeye.inventory.schemas import InventoryResult

    rows = _result_rows(InventoryResult(
        result_kind="circular", document_id="c-1", title="A Circular",
        reference="BPRD 1 of 2020", matched_via=["lexical"], evidence=[],
    ))

    assert len(rows) == 1
    assert rows[0][0] == "BPRD 1 of 2020"
    assert len(rows[0]) == len(INVENTORY_HEADERS)


def test_strict_coverage_refuses_an_incomplete_index(indexed):
    db, collection = indexed
    for chunk_id in [i for i in collection.records if i.startswith("c-audit__")]:
        collection.delete([chunk_id])
    ledger.reconcile(db, collection, FakeConfig(), write=True)

    request = InventorySearchRequest(
        query="AML", generate_terms=False, use_hyde=False, skip_adjudication=True,
        require_complete_coverage=True,
    )

    with pytest.raises(Exception) as excinfo:
        make_service(collection).search(request, db)
    assert "semantic_index_incomplete" in getattr(excinfo.value, "code", "")


def test_law_results_carry_version_and_hierarchy(indexed):
    db, collection = indexed
    request = InventorySearchRequest(
        query="anti-money laundering", sources=["laws"], generate_terms=False,
        use_hyde=False, skip_adjudication=True, extract_spans=False,
    )

    response = make_service(collection).search(request, db)

    law = next(r for r in response.results if r.document_id == "l-aml")
    assert law.result_kind == "law" and law.version_id == "l-aml-v1"


def test_orphan_chunks_cannot_produce_a_hit(indexed):
    """Gap 0.2.1: a chunk with no current ledger entry must not be scored.

    Otherwise a deleted circular's leftovers keep surfacing as semantic matches long
    after the database stopped considering them part of the corpus.
    """
    db, collection = indexed
    collection.add(
        documents=["Ghost circular about anti-money laundering and AML controls."],
        embeddings=FakeBackend().embed_documents(["aml aml aml money laundering"]),
        ids=["c-ghost__chunk_0"],
        metadatas=[{"circular_id": "c-ghost", "doc_type": "circular", "ref": "Chunk 1"}],
    )
    SNAPSHOT_CACHE.clear()

    snapshot = load_snapshot(collection, "snap", ledger.indexed_source_ids(db))

    assert snapshot.excluded_orphans == 1
    assert all("c-ghost" not in cid for cid in snapshot.chunk_ids)


def test_orphan_exclusion_is_reported_in_coverage(indexed):
    db, collection = indexed
    collection.add(
        documents=["Ghost."], embeddings=[[0.0, 0.0, 1.0]], ids=["c-ghost__chunk_0"],
        metadatas=[{"circular_id": "c-ghost", "doc_type": "circular"}],
    )
    SNAPSHOT_CACHE.clear()
    request = InventorySearchRequest(
        query="AML", generate_terms=False, use_hyde=False, skip_adjudication=True,
        extract_spans=False,
    )

    response = make_service(collection).search(request, db)

    assert any("excluded from scoring" in w for w in response.coverage.warnings)
    assert all(r.document_id != "c-ghost" for r in response.results)


def test_search_aborts_when_the_embedding_model_changed(indexed):
    """Gap 0.2.2: cosine across two embedding spaces is a meaningless number."""
    db, collection = indexed

    class OtherModel:
        provider = "fastembed"
        model = "a-different-model"

    service = InventorySearchService(collection, FakeBackend(), OtherModel(), llm=None)
    request = InventorySearchRequest(
        query="AML", generate_terms=False, use_hyde=False, skip_adjudication=True,
    )

    with pytest.raises(EmbeddingFingerprintMismatch):
        service.search(request, db)


def test_search_aborts_on_a_dimension_mismatch(indexed):
    db, collection = indexed

    class WiderBackend(FakeBackend):
        def embed_queries(self, queries):
            return [[1.0, 0.0, 0.0, 0.0] for _ in queries]

    SNAPSHOT_CACHE.clear()
    service = InventorySearchService(collection, WiderBackend(), FakeConfig(), llm=None)
    request = InventorySearchRequest(
        query="AML", generate_terms=False, use_hyde=False, skip_adjudication=True,
    )

    with pytest.raises(EmbeddingFingerprintMismatch) as excinfo:
        service.search(request, db)
    assert "dimensional" in str(excinfo.value)


def test_a_source_that_chunks_to_nothing_is_empty_not_stale(corpus_db):
    """Page markers with no words are an empty source, not an indexing failure.

    Marking them stale would make `--repair` re-embed them on every run forever.
    """
    corpus_db.add(Attachment(
        id="a-markers", circular_id="c-aml", filename="scan.pdf",
        original_url="https://example.test/s", file_type="pdf",
        content_text="[[SBPEYE_PAGE:1]]\n[[SBPEYE_PAGE:2]]\n", extraction_status="ok",
    ))
    corpus_db.commit()

    ledger.reconcile(corpus_db, FakeCollection(), FakeConfig(), write=True)
    row = corpus_db.query(SemanticIndexSource).filter_by(source_id="a-markers").one()

    assert row.status == corpus.STATUS_EMPTY
    assert row.expected_chunks == 0


def test_indexing_a_law_writes_its_ledger_row(corpus_db, monkeypatch):
    """Gap 1: the ledger must track live indexing, not only batch reconciliation.

    Without this, coverage keeps asserting a completeness it last checked at the previous
    audit, which is exactly the claim the ledger exists to make honest.
    """
    import sbpeye.scraper.circulars as circulars_mod

    fake = FakeCollection()
    monkeypatch.setattr(circulars_mod, "collection", fake)
    monkeypatch.setattr(circulars_mod, "embedding_backend", FakeBackend())

    from sbpeye.scraper.laws import vectorize_law_document

    document = corpus_db.query(RegDocument).filter_by(id="l-aml").one()
    count = vectorize_law_document(corpus_db, document)

    row = corpus_db.query(SemanticIndexSource).filter_by(source_id="l-aml-v1").one()
    assert row.status == corpus.STATUS_INDEXED
    assert row.indexed_chunks == row.expected_chunks == count
    assert row.logical_document_id == "l-aml"
    assert row.version_id == "l-aml-v1"


def test_a_failed_index_write_is_recorded_as_index_error(corpus_db, monkeypatch):
    """A Chroma failure must leave evidence, not an absent row that reads as 'fine'."""
    import sbpeye.scraper.circulars as circulars_mod

    class ExplodingCollection(FakeCollection):
        def add(self, *args, **kwargs):
            raise RuntimeError("chroma down")

    monkeypatch.setattr(circulars_mod, "collection", ExplodingCollection())
    monkeypatch.setattr(circulars_mod, "embedding_backend", FakeBackend())

    circular = corpus_db.query(Circular).filter_by(id="c-aml").one()
    circulars_mod._index_circular(circular, db=corpus_db)

    row = corpus_db.query(SemanticIndexSource).filter_by(source_id="c-aml").one()
    assert row.status == corpus.STATUS_INDEX_ERROR
    assert "chroma down" in (row.error or "")
    assert row.indexed_chunks == 0


def test_ledger_write_failure_does_not_break_indexing(corpus_db, monkeypatch):
    """Bookkeeping must never fail an index write; the row simply stays absent."""
    import sbpeye.scraper.circulars as circulars_mod
    from sbpeye.inventory import ledger as ledger_mod

    monkeypatch.setattr(circulars_mod, "collection", FakeCollection())
    monkeypatch.setattr(circulars_mod, "embedding_backend", FakeBackend())
    monkeypatch.setattr(
        ledger_mod, "expected_chunk_count",
        lambda text: (_ for _ in ()).throw(RuntimeError("ledger boom")),
    )

    circular = corpus_db.query(Circular).filter_by(id="c-aml").one()
    circulars_mod._index_circular(circular, db=corpus_db)  # must not raise

    assert corpus_db.query(SemanticIndexSource).filter_by(source_id="c-aml").count() == 0


def test_filters_apply_before_retrieval_so_counts_describe_the_search(indexed):
    """Gap 2: filtering afterwards leaves the candidate counts describing excluded docs."""
    db, collection = indexed
    request = InventorySearchRequest(
        query="AML", sources=["circulars"], generate_terms=False, use_hyde=False,
        skip_adjudication=True, extract_spans=False, semantic_band=100,
        filters=InventoryFilters(departments=["BPRD"]),
    )

    response = make_service(collection).search(request, db)

    # c-mention is BSD, so it must not be counted by either arm.
    assert all(r.department == "BPRD" for r in response.results)
    assert response.coverage.candidates_semantic <= 2
    assert all(r.document_id != "c-mention" for r in response.results)


def test_a_filtered_out_document_is_not_even_a_lexical_candidate(indexed):
    db, collection = indexed
    request = InventorySearchRequest(
        query="AML", sources=["circulars"], generate_terms=False, use_hyde=False,
        skip_adjudication=True, extract_spans=False, semantic_band=0,
        filters=InventoryFilters(departments=["BPRD"]),
    )

    response = make_service(collection).search(request, db)

    # Both c-aml (BPRD) and c-mention (BSD) contain "AML"; only the BPRD one counts.
    assert response.coverage.candidates_lexical == 1


def test_coverage_reports_lexical_and_vector_arms_separately(indexed):
    """Gap 1: the arms fail independently, so one number cannot describe both."""
    db, collection = indexed
    db.execute(text("DELETE FROM circulars_fts WHERE circular_id = 'c-audit'"))
    db.commit()

    request = InventorySearchRequest(
        query="AML", generate_terms=False, use_hyde=False, skip_adjudication=True,
        extract_spans=False,
    )
    response = make_service(collection).search(request, db)

    coverage = response.coverage
    assert coverage.lexical_gaps == 1
    # The vector arm is untouched: the document is still semantically reachable.
    assert coverage.stale_or_missing_index == 0
    assert coverage.is_complete is False
    assert any("no FTS row" in w for w in coverage.warnings)


def test_max_results_truncates_explicitly(indexed):
    """Gap 3: a capped inventory must say so rather than look complete."""
    db, collection = indexed
    request = InventorySearchRequest(
        query="AML", generate_terms=False, use_hyde=False, skip_adjudication=True,
        extract_spans=False, max_results=1,
    )

    response = make_service(collection).search(request, db)

    assert len(response.results) == 1
    assert response.results_truncated >= 1
    assert response.truncated is True
    assert any("max_results" in w for w in response.coverage.warnings)


def test_blank_query_is_rejected(indexed):
    db, collection = indexed

    with pytest.raises(InvalidQuery):
        make_service(collection).search(InventorySearchRequest(query="   "), db)


def test_response_serializes_without_law_keys_on_circulars(indexed):
    db, collection = indexed
    request = InventorySearchRequest(
        query="AML", generate_terms=False, use_hyde=False, skip_adjudication=True,
        extract_spans=False,
    )

    payload = make_service(collection).search(request, db).to_dict()

    circular_rows = [r for r in payload["results"] if r["result_kind"] == "circular"]
    assert circular_rows and "version_id" not in circular_rows[0]
