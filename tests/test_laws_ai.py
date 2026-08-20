"""Phase B of docs/LAWS_AI_PLAN.md: a law reaches the existing extractors.

The corpus builder decides what gets analysed and, just as importantly, records why
anything did not — 33 of the 133 rows in the live corpus have no analysable text, and each
has a different reason a reader deserves to be told.
"""

import json
import re
from datetime import datetime
from types import SimpleNamespace

import pytest
from sqlalchemy.pool import StaticPool

from sbpeye.ai import AIClient, AIConfig
from sbpeye.laws_ai import (
    GAP_CIRCULAR_BACKED,
    GAP_EXTERNAL,
    GAP_MANIFEST,
    GAP_MISSING_FILE,
    GAP_MISSING_TEXT,
    GAP_NO_CURRENT_VERSION,
    GAP_UNSUPPORTED_FILE_TYPE,
    STRUCTURAL_GAPS,
    law_corpus,
)
from sbpeye.laws_links import law_label
from sbpeye.models import RegDocument as RegDocumentForTest
from sbpeye.scraper.laws import law_document

from test_reg_models import make_document, make_session, make_version


def build(session, *, version_overrides=None, **document_overrides):
    """A document with one current version, persisted so relationships resolve."""
    document = make_document(**document_overrides)
    session.add(document)
    if version_overrides is not None:
        version = make_version(
            "ver-1", document.id, is_current=1, **version_overrides
        )
        session.add(version)
    session.flush()
    return document


def test_corpus_carries_the_archived_pdf_so_docling_reads_the_original(tmp_path):
    session = make_session()
    pdf = tmp_path / "sme-prs.pdf"
    pdf.write_bytes(b"%PDF placeholder")
    document = build(
        session,
        version_overrides={
            "file_type": "pdf",
            "local_path": str(pdf),
            "content_text": "1. Banks shall maintain capital.",
        },
    )

    documents, gaps = law_corpus(document)

    assert gaps == []
    assert len(documents) == 1
    # The file, not the re-serialised text: without this the parse loses page numbers and
    # every citation in the resulting checklist becomes unverifiable.
    assert documents[0]["local_path"] == str(pdf)
    assert documents[0]["doc_type"] == "law"


def test_an_html_law_has_no_file_and_falls_back_to_its_text():
    session = make_session()
    document = build(
        session,
        version_overrides={
            "file_type": "html",
            "local_path": None,
            "content_text": "1. Banks shall report monthly.",
        },
    )

    documents, gaps = law_corpus(document)

    assert gaps == []
    assert "local_path" not in documents[0]
    assert documents[0]["text"] == "1. Banks shall report monthly."


def test_a_part_is_named_with_its_container():
    session = make_session()
    parent = build(session, document_id="fe-manual", title="Foreign Exchange Manual")
    child = build(
        session,
        document_id="fe-ch-12",
        title="EXPORTS",
        parent_id=parent.id,
        part_label="Chapter 12",
        version_overrides={"file_type": "html", "local_path": None, "content_text": "Rules."},
    )

    documents, _ = law_corpus(child)

    assert documents[0]["doc_label"] == "Foreign Exchange Manual - Chapter 12: EXPORTS"
    assert law_label(child) == documents[0]["doc_label"]


def test_a_container_is_a_manifest_not_an_empty_document():
    """The FE Manual holds no bytes of its own; its 26 chapters are analysed separately."""
    session = make_session()
    document = build(
        session,
        version_overrides={"file_type": "manifest", "local_path": None, "content_text": None},
    )

    documents, gaps = law_corpus(document)

    assert documents == []
    assert gaps[0]["reason"] == GAP_MANIFEST
    assert gaps[0]["doc_type"] == "law"


def test_a_spreadsheet_is_reported_as_unreadable_rather_than_crashing_docling():
    session = make_session()
    document = build(
        session,
        title="Questionnaire",
        version_overrides={"file_type": "xls", "local_path": None, "content_text": None},
    )

    documents, gaps = law_corpus(document)

    assert documents == []
    assert gaps[0]["reason"] == GAP_UNSUPPORTED_FILE_TYPE


def test_a_circular_backed_row_defers_to_the_circular_corpus():
    session = make_session()
    document = build(session, circular_id="circular-1")

    documents, gaps = law_corpus(document)

    assert documents == []
    assert gaps[0]["reason"] == GAP_CIRCULAR_BACKED


def test_an_external_law_and_a_stub_are_distinguished():
    session = make_session()
    external = build(session, document_id="ext-1", is_external=1)
    stub = build(session, document_id="stub-1")

    assert law_corpus(external)[1][0]["reason"] == GAP_EXTERNAL
    assert law_corpus(stub)[1][0]["reason"] == GAP_NO_CURRENT_VERSION


def test_a_missing_archive_file_is_a_gap_not_a_silent_text_fallback(tmp_path):
    """Falling back would produce a parse with no pages and no sign anything was wrong."""
    session = make_session()
    document = build(
        session,
        version_overrides={
            "file_type": "pdf",
            "local_path": str(tmp_path / "gone.pdf"),
            "content_text": "1. Banks shall maintain capital.",
        },
    )

    documents, gaps = law_corpus(document)

    assert documents == []
    assert gaps[0]["reason"] == GAP_MISSING_FILE


def test_an_unextractable_pdf_reports_the_extraction_error():
    session = make_session()
    document = build(
        session,
        version_overrides={
            "file_type": "pdf",
            "local_path": None,
            "content_text": "",
            "extraction_error": "scanned, no text layer",
        },
    )

    documents, gaps = law_corpus(document)

    assert documents == []
    assert gaps[0]["reason"] == GAP_MISSING_TEXT
    assert gaps[0]["error"] == "scanned, no text layer"


def test_structural_gaps_are_the_ones_no_rerun_can_fix():
    assert GAP_MANIFEST in STRUCTURAL_GAPS
    assert GAP_CIRCULAR_BACKED in STRUCTURAL_GAPS
    # These two can change: SBP can serve the file, or extraction can be improved.
    assert GAP_MISSING_FILE not in STRUCTURAL_GAPS
    assert GAP_MISSING_TEXT not in STRUCTURAL_GAPS


def test_superseded_editions_are_never_analysed():
    """Obligations quoted from wording SBP no longer publishes are worse than none."""
    session = make_session()
    document = build(
        session,
        version_overrides={"file_type": "html", "local_path": None, "content_text": "New text."},
    )
    session.add(make_version(
        "ver-old", document.id, is_current=0, file_type="html",
        local_path=None, content_text="Superseded text.",
        first_seen_at=datetime(2024, 1, 1),
    ))
    session.flush()

    documents, _ = law_corpus(document)

    assert [doc["text"] for doc in documents] == ["New text."]


# --- the extractors, driven without a Circular --------------------------------------


def test_generate_checklist_runs_on_a_law_corpus(monkeypatch):
    session = make_session()
    document = build(
        session,
        title="Prudential Regulations for Housing Finance",
        version_overrides={
            "file_type": "html",
            "local_path": None,
            "content_text": (
                "# CAPITAL\n\n1. Banks shall maintain capital against housing finance."
            ),
        },
    )
    documents, gaps = law_corpus(document)

    client = AIClient(AIConfig())
    prompts = []

    def complete(system, user, **kwargs):
        prompts.append(user)
        source_ids = re.findall(r"\[SOURCE_ID: ([^]]+)]", user)
        return json.dumps({"items": [{
            "requirement": "Banks must maintain capital against housing finance.",
            "classification": "required",
            "source_unit_ids": [source_ids[0]],
        }]})

    monkeypatch.setattr(client, "_complete", complete)
    monkeypatch.setattr(client, "resolve_context_budget", lambda: 100_000)

    result = client.generate_checklist(
        label=document.title, documents=documents, gaps=gaps
    )

    assert result["status"] == "completed"
    assert result["checklist_items"][0]["classification"] == "required"
    # The model is told what it is reading, and it is not told it is a circular.
    assert f"Subject: {document.title}" in prompts[0]
    assert "Circular:" not in prompts[0]


def test_extract_entities_runs_on_a_law_corpus(monkeypatch):
    session = make_session()
    document = build(
        session,
        version_overrides={
            "file_type": "html",
            "local_path": None,
            "content_text": "# CAPITAL\n\n1. Banks shall maintain a CAR of 11.5%.",
        },
    )
    documents, _ = law_corpus(document)

    client = AIClient(AIConfig())

    def complete(system, user, **kwargs):
        source_ids = re.findall(r"\[SOURCE_ID: ([^]]+)]", user)
        return json.dumps({"entities": [{
            "entity_type": "ratio", "metric": "CAR", "comparator": "min",
            "value_numeric": 11.5, "value_high": None, "unit": "%",
            "value_text": "11.5%", "subject": "banks", "effective_date": None,
            "context_snippet": "CAR of 11.5%", "source_unit_ids": [source_ids[0]],
            "confidence": 0.9,
        }]})

    monkeypatch.setattr(client, "_complete", complete)
    monkeypatch.setattr(client, "resolve_context_budget", lambda: 100_000)

    entities = client.extract_entities(label=document.title, documents=documents)

    assert [entity["metric"] for entity in entities] == ["CAR"]
    assert entities[0]["source_unit_id"] is not None


def test_a_circular_still_names_itself_without_an_explicit_label():
    circular = SimpleNamespace(
        reference="BPRD Circular No. 1 of 2025", title="Reporting requirements"
    )
    assert AIClient._subject_label(circular, None) == "BPRD Circular No. 1 of 2025"
    assert AIClient._subject_label(circular, "Override") == "Override"
    assert AIClient._subject_label(None, None) == ""


def test_law_document_omits_local_path_for_a_non_pdf():
    session = make_session()
    document = build(session, version_overrides={"file_type": "xls", "local_path": "/tmp/x.xls"})

    payload = law_document(document, document.current_version)

    # Docling registers PDF and Markdown only; handing it a spreadsheet path raises.
    assert "local_path" not in payload


# --- phase C: storage, jobs, endpoint -----------------------------------------------


class FakeLawClient:
    """Records what the generation pipeline asked for, without a provider."""

    def __init__(self):
        self.config = SimpleNamespace(max_context_tokens=1_000_000)
        self.summarize_calls = []
        self.reduce_calls = []
        self.tag_basis = []

    def resolve_context_budget(self):
        return 100_000

    def summarize(self, title, content_text, *, subject="circular"):
        self.summarize_calls.append((title, content_text, subject))
        return f"Summary of {title}."

    def reduce_summaries(self, title, summaries, *, subject="document"):
        self.reduce_calls.append(summaries)
        return f"Reduced summary of {title}."

    def generate_tags(self, title, content_text, *, subject="circular"):
        self.tag_basis.append(content_text)
        return ["Capital Adequacy"]

    def generate_checklist(self, circular=None, *, label=None, documents=None,
                           gaps=None, progress_callback=None, **kwargs):
        if progress_callback:
            progress_callback(1, 1)
        return {"status": "completed", "checklist_items": [{"requirement": "Hold capital."}]}

    def extract_entities(self, circular=None, *, label=None, documents=None,
                         gaps=None, progress_callback=None, **kwargs):
        if progress_callback:
            progress_callback(1, 1)
        return [{
            "entity_type": "ratio", "metric": "CAR", "comparator": "min",
            "value_numeric": 11.5, "value_high": None, "unit": "%",
            "value_text": "11.5%", "subject": "banks", "effective_date": None,
            "context_snippet": "CAR of 11.5%", "source_unit_id": "u1",
            "page_start": 3, "confidence": 0.9,
        }]


def test_analysis_is_stored_against_the_edition_that_produced_it(monkeypatch):
    """§4: a version's bytes never change, so its analysis never needs invalidating."""
    from sqlalchemy.orm import sessionmaker

    import sbpeye.laws_ai as laws_ai_module
    from sbpeye.models import AIGenerationJob, RegDocument
    from test_laws_search import add_law, make_session

    db, engine = make_session()
    document, version = add_law(db, "doc-1", file_type="html", index=False)
    factory = sessionmaker(bind=engine, autoflush=False)
    db.add(AIGenerationJob(
        id="job-1", target_kind="law", document_id="doc-1", feature="all",
        status="queued", created_at=datetime(2026, 8, 14),
    ))
    db.commit()

    client = FakeLawClient()
    monkeypatch.setattr(laws_ai_module, "SessionLocal", factory)
    monkeypatch.setattr(laws_ai_module, "get_ai_client", lambda db: client)

    laws_ai_module.run_law_generation_job("job-1")

    db.expire_all()
    job = db.query(AIGenerationJob).filter(AIGenerationJob.id == "job-1").one()
    stored = db.query(RegDocument).filter(RegDocument.id == "doc-1").one()
    assert job.status == "succeeded", job.error
    assert stored.current_version.summary == "Summary of Prudential Regulations for SME Financing."
    assert json.loads(stored.current_version.tags) == ["Capital Adequacy"]
    # The document row stays clean: it is reserved for container rollups.
    assert stored.summary is None

    # A law's values carry the edition they were read out of, so a later capture can be
    # compared against this one rather than overwriting it.
    value = stored.entities[0]
    assert (value.subject_kind, value.metric, value.circular_id) == ("law", "CAR", None)
    assert value.version_id == stored.current_version.id
    assert stored.current_version.entities_generated_at is not None


def test_tags_are_drawn_from_the_summary_not_the_head_of_the_text(monkeypatch):
    """Tags taken from the first window of a 200-page manual describe its preface."""
    from sqlalchemy.orm import sessionmaker

    import sbpeye.laws_ai as laws_ai_module
    from sbpeye.models import AIGenerationJob
    from test_laws_search import add_law, make_session

    db, engine = make_session()
    add_law(db, "doc-1", file_type="html", index=False)
    db.add(AIGenerationJob(
        id="job-1", target_kind="law", document_id="doc-1", feature="all",
        status="queued", created_at=datetime(2026, 8, 14),
    ))
    db.commit()

    client = FakeLawClient()
    monkeypatch.setattr(laws_ai_module, "SessionLocal", sessionmaker(bind=engine, autoflush=False))
    monkeypatch.setattr(laws_ai_module, "get_ai_client", lambda db: client)

    laws_ai_module.run_law_generation_job("job-1")

    assert client.tag_basis == ["Summary of Prudential Regulations for SME Financing."]
    # And the model is told what kind of instrument it is reading.
    assert client.summarize_calls[0][2] == "regulation"


def test_summarize_law_is_one_call_when_the_document_fits():
    from sbpeye.laws_ai import summarize_law

    client = FakeLawClient()
    summary = summarize_law(client, "SME PRs", "Short text.", "regulation")

    assert summary == "Summary of SME PRs."
    assert len(client.summarize_calls) == 1
    assert client.reduce_calls == []


def test_summarize_law_maps_and_reduces_a_document_too_long_for_one_call():
    from sbpeye.laws_ai import summarize_law

    client = FakeLawClient()
    client.resolve_context_budget = lambda: 1_000  # 4,000 chars per window
    # 20k chars: a small law by the corpus's standards, and already five windows.
    text = "\n\n".join(f"Paragraph {index}. " + "x" * 900 for index in range(22))

    summary = summarize_law(client, "FE Manual", text, "law")

    assert len(client.summarize_calls) == 6
    assert summary == "Reduced summary of FE Manual."
    assert client.reduce_calls[0] == ["Summary of FE Manual."] * 6
    # Windows break on paragraph boundaries, so no window splits a sentence.
    assert all(window.startswith("Paragraph") for _, window, _ in client.summarize_calls)


def test_a_superseded_editions_summary_is_never_shown_as_the_documents():
    """The §4 invariant: un-analysed in force reads as un-analysed, not as the old answer."""
    from sbpeye.api.serializers import _law_detail
    from test_laws_search import add_law, make_session

    db, _ = make_session()
    document, current = add_law(db, "doc-1", file_type="html", index=False)
    superseded = make_version(
        "doc-1-old", "doc-1", is_current=0, file_type="html",
        content_text="Old text.", summary="Summary of the 2024 edition.",
        summary_generated_at=datetime(2024, 1, 1),
        first_seen_at=datetime(2024, 1, 1),
    )
    db.add(superseded)
    db.commit()

    payload = _law_detail(document)

    assert payload["summary"] is None
    assert payload["generation"]["summary"] is None
    assert payload["checklist_available"] is False


# --- the endpoint -------------------------------------------------------------------


@pytest.fixture
def law_client(monkeypatch):
    """TestClient over an isolated DB, with the background worker stubbed.

    Starlette runs background tasks synchronously once the response is returned, so an
    unstubbed worker would reach a real provider from inside a route test.
    """
    from fastapi.testclient import TestClient
    from sqlalchemy.orm import sessionmaker

    import sbpeye.auth_routes as auth_routes_module
    import sbpeye.main as main_module
    from sbpeye.database import get_db
    from test_laws_search import EmptyCollection, make_session

    db, engine = make_session()
    factory = sessionmaker(bind=engine, autoflush=False)

    def override_get_db():
        session = factory()
        try:
            yield session
        finally:
            session.close()

    started: list[str] = []
    monkeypatch.setattr(main_module, "SessionLocal", factory)
    monkeypatch.setattr(main_module, "_warm_up_search_index", lambda: None)
    monkeypatch.setattr(main_module, "run_law_generation_job", started.append)
    monkeypatch.setattr("sbpeye.search.collection", EmptyCollection())
    main_module.app.dependency_overrides[get_db] = override_get_db
    # `resolve_request_user` opens its own session, so it has to be pointed at this DB
    # too or the middleware looks the signed-in user up in the developer's real one.
    monkeypatch.setattr(auth_routes_module, "AppSessionLocal", factory)
    with TestClient(main_module.app) as test_client:
        # This fixture builds its own client rather than using conftest's, so it has to
        # establish a session itself; every route sits behind the auth middleware.
        from conftest import sign_in

        sign_in(test_client, factory, is_admin=True)
        yield test_client, db, started
    main_module.app.dependency_overrides.clear()


def test_generate_endpoint_queues_a_law_job(law_client):
    from sbpeye.models import AIGenerationJob
    from test_laws_search import add_law

    test_client, db, started = law_client
    add_law(db, "doc-1", file_type="html", index=False)

    response = test_client.post("/api/laws/doc-1/generate", json={"feature": "summary"})

    assert response.status_code == 202
    payload = response.json()
    assert payload["target_kind"] == "law"
    assert payload["document_id"] == "doc-1"
    assert payload["circular_id"] is None
    assert started == [payload["id"]]
    assert db.query(AIGenerationJob).filter(AIGenerationJob.id == payload["id"]).one()


def test_generate_endpoint_explains_why_a_container_cannot_be_analysed(law_client):
    """"No content" is not actionable; "analyse its parts" is."""
    test_client, db, _ = law_client
    from test_laws_search import add_law

    add_law(db, "fe-manual", file_type="manifest", text_body=None, index=False)

    response = test_client.post("/api/laws/fe-manual/generate", json={"feature": "all"})

    assert response.status_code == 422
    payload = response.json()
    assert payload["reason"] == GAP_MANIFEST
    assert payload["structural"] is True
    assert "parts" in payload["error"]


def test_generate_endpoint_marks_a_recoverable_gap_as_non_structural(law_client):
    test_client, db, _ = law_client
    from test_laws_search import add_law

    add_law(db, "stub-1", file_type="pdf", text_body="", index=False)

    payload = test_client.post(
        "/api/laws/stub-1/generate", json={"feature": "summary"}
    ).json()

    # SBP may serve the file tomorrow, so the UI should not present this as permanent.
    assert payload["structural"] is False


def test_generate_endpoint_rejects_consolidation_which_laws_do_not_have(law_client):
    test_client, db, _ = law_client
    from test_laws_search import add_law

    add_law(db, "doc-1", file_type="html", index=False)

    response = test_client.post(
        "/api/laws/doc-1/generate", json={"feature": "consolidation"}
    )

    assert response.status_code == 400
    assert "summary" in response.json()["error"]


def test_generate_endpoint_reports_an_in_flight_job(law_client):
    from sbpeye.models import AIGenerationJob
    from test_laws_search import add_law

    test_client, db, _ = law_client
    add_law(db, "doc-1", file_type="html", index=False)
    db.add(AIGenerationJob(
        id="job-1", target_kind="law", document_id="doc-1", feature="all",
        status="running", created_at=datetime(2026, 8, 14),
    ))
    db.commit()

    response = test_client.post("/api/laws/doc-1/generate", json={"feature": "summary"})

    assert response.status_code == 409
    assert response.json()["job"]["id"] == "job-1"


def test_generate_endpoint_404s_for_an_unknown_document(law_client):
    test_client, _, _ = law_client
    assert test_client.post(
        "/api/laws/nope/generate", json={"feature": "summary"}
    ).status_code == 404


def test_law_detail_exposes_generation_state(law_client):
    from test_laws_search import add_law

    test_client, db, _ = law_client
    _, version = add_law(db, "doc-1", file_type="html", index=False)
    version.summary = "A summary of the edition in force."
    version.summary_generated_at = datetime(2026, 8, 14)
    db.commit()

    payload = test_client.get("/api/laws/doc-1").json()

    assert payload["summary"] == "A summary of the edition in force."
    assert payload["generation"]["summary"] is not None
    assert payload["generation"]["checklist"] is None
    assert payload["checklist_available"] is False


def test_the_job_table_migrates_to_hold_a_law_job():
    """The one migration here that is not an ALTER: SQLite cannot relax NOT NULL."""
    from sqlalchemy import create_engine, inspect, text
    from sbpeye.database import _ensure_columns

    engine = create_engine("sqlite://", poolclass=StaticPool)
    with engine.begin() as conn:
        conn.execute(text(
            "CREATE TABLE ai_generation_jobs ("
            "id VARCHAR NOT NULL PRIMARY KEY, circular_id VARCHAR NOT NULL, "
            "feature VARCHAR NOT NULL, status VARCHAR NOT NULL, error TEXT, "
            "created_at DATETIME NOT NULL, started_at DATETIME, completed_at DATETIME)"
        ))
        conn.execute(text(
            "INSERT INTO ai_generation_jobs (id, circular_id, feature, status, created_at) "
            "VALUES ('old-1', 'circ-1', 'summary', 'succeeded', CURRENT_TIMESTAMP)"
        ))

    _ensure_columns(bind=engine)

    columns = {c["name"]: c for c in inspect(engine).get_columns("ai_generation_jobs")}
    assert columns["circular_id"]["nullable"]
    with engine.begin() as conn:
        # The work log survives the rebuild, tagged as what it always was.
        assert conn.execute(text(
            "SELECT target_kind FROM ai_generation_jobs WHERE id = 'old-1'"
        )).scalar() == "circular"
        conn.execute(text(
            "INSERT INTO ai_generation_jobs (id, target_kind, document_id, feature, status, created_at) "
            "VALUES ('new-1', 'law', 'doc-1', 'summary', 'queued', CURRENT_TIMESTAMP)"
        ))


# --- phase D: regulatory values across both corpora ---------------------------------


def add_value(db, **overrides):
    from sbpeye.models import CircularEntity

    fields = dict(
        subject_kind="law", entity_type="ratio", metric="CAR", comparator="min",
        value_numeric=11.5, unit="%", value_text="11.5%", subject="banks",
        created_at=datetime(2026, 8, 14),
    )
    fields.update(overrides)
    entity = CircularEntity(**fields)
    db.add(entity)
    db.commit()
    return entity


def test_the_values_query_returns_both_corpora(law_client):
    """The inner join this replaced dropped every law-sourced value silently."""
    from sbpeye.models import Circular
    from test_laws_search import add_law

    test_client, db, _ = law_client
    _, version = add_law(db, "doc-1", file_type="html", index=False)
    db.add(Circular(id="circ-1", title="Capital circular", reference="BPRD 1",
                    date=datetime(2025, 1, 1), status="active"))
    db.commit()
    add_value(db, subject_kind="law", document_id="doc-1", version_id=version.id)
    add_value(db, subject_kind="circular", circular_id="circ-1", metric="LCR")

    payload = test_client.get("/api/circulars/entities/query").json()

    assert payload["total"] == 2
    kinds = {row["subject_kind"] for row in payload["results"]}
    assert kinds == {"law", "circular"}

    law_row = next(r for r in payload["results"] if r["subject_kind"] == "law")
    assert law_row["circular_id"] is None
    assert law_row["document"]["display_title"] == "Prudential Regulations for SME Financing"
    assert law_row["document"]["in_force"] is True


def test_the_values_query_can_be_narrowed_to_one_corpus(law_client):
    from sbpeye.models import Circular
    from test_laws_search import add_law

    test_client, db, _ = law_client
    _, version = add_law(db, "doc-1", file_type="html", index=False)
    db.add(Circular(id="circ-1", title="Capital circular", reference="BPRD 1",
                    date=datetime(2025, 1, 1), status="active"))
    db.commit()
    add_value(db, document_id="doc-1", version_id=version.id)
    add_value(db, subject_kind="circular", circular_id="circ-1", metric="LCR")

    laws = test_client.get("/api/circulars/entities/query", params={"source": "laws"}).json()
    circulars = test_client.get(
        "/api/circulars/entities/query", params={"source": "circulars"}
    ).json()

    assert [r["metric"] for r in laws["results"]] == ["CAR"]
    assert [r["metric"] for r in circulars["results"]] == ["LCR"]


def test_current_only_ages_out_a_superseded_edition_not_just_a_superseded_circular(law_client):
    """Each corpus decides currency differently; expressing only one filters out the other."""
    from sbpeye.models import Circular
    from test_laws_search import add_law

    test_client, db, _ = law_client
    _, current = add_law(db, "doc-1", file_type="html", index=False)
    old = make_version("doc-1-old", "doc-1", is_current=0, file_type="html",
                       content_text="Old.", first_seen_at=datetime(2024, 1, 1))
    db.add(old)
    db.add(Circular(id="circ-1", title="Superseded circular", reference="BPRD 1",
                    date=datetime(2025, 1, 1), status="superseded"))
    db.commit()
    add_value(db, document_id="doc-1", version_id=current.id, metric="CAR")
    add_value(db, document_id="doc-1", version_id=old.id, metric="Old CAR")
    add_value(db, subject_kind="circular", circular_id="circ-1", metric="Stale LCR")

    payload = test_client.get(
        "/api/circulars/entities/query", params={"current_only": True}
    ).json()

    assert [row["metric"] for row in payload["results"]] == ["CAR"]


def test_a_law_detail_shows_only_the_values_of_the_edition_in_force(law_client):
    from test_laws_search import add_law

    test_client, db, _ = law_client
    _, current = add_law(db, "doc-1", file_type="html", index=False)
    old = make_version("doc-1-old", "doc-1", is_current=0, file_type="html",
                       content_text="Old.", first_seen_at=datetime(2024, 1, 1))
    db.add(old)
    db.commit()
    add_value(db, document_id="doc-1", version_id=current.id, metric="CAR")
    add_value(db, document_id="doc-1", version_id=old.id, metric="CAR (2024)")

    payload = test_client.get("/api/laws/doc-1").json()

    assert [entity["metric"] for entity in payload["entities"]] == ["CAR"]


def test_the_entities_table_migrates_to_hold_a_law_value():
    from sqlalchemy import create_engine, inspect, text
    from sbpeye.database import _ensure_columns

    engine = create_engine("sqlite://", poolclass=StaticPool)
    with engine.begin() as conn:
        conn.execute(text(
            "CREATE TABLE circular_entities ("
            "id INTEGER NOT NULL PRIMARY KEY, circular_id VARCHAR NOT NULL, "
            "entity_type VARCHAR NOT NULL, metric VARCHAR, comparator VARCHAR, "
            "value_numeric FLOAT, value_high FLOAT, unit VARCHAR, value_text VARCHAR, "
            "subject VARCHAR, effective_date DATETIME, context_snippet TEXT, "
            "source_unit_id VARCHAR, page_start INTEGER, confidence FLOAT, "
            "created_at DATETIME NOT NULL)"
        ))
        conn.execute(text(
            "INSERT INTO circular_entities (id, circular_id, entity_type, metric, created_at) "
            "VALUES (7, 'circ-1', 'ratio', 'CAR', CURRENT_TIMESTAMP)"
        ))

    _ensure_columns(bind=engine)

    columns = {c["name"]: c for c in inspect(engine).get_columns("circular_entities")}
    assert columns["circular_id"]["nullable"]
    with engine.begin() as conn:
        # The id is preserved, not reassigned by the rebuild.
        row = conn.execute(text(
            "SELECT id, subject_kind FROM circular_entities WHERE metric = 'CAR'"
        )).one()
        assert row == (7, "circular")
        conn.execute(text(
            "INSERT INTO circular_entities "
            "(subject_kind, document_id, version_id, entity_type, metric, created_at) "
            "VALUES ('law', 'doc-1', 'ver-1', 'ratio', 'LCR', CURRENT_TIMESTAMP)"
        ))


# --- phase E: relationships ---------------------------------------------------------


def test_a_laws_own_text_yields_candidates_with_the_wording_that_names_them():
    """Deterministic: the name is in the text. What it *means* is the AI pass's job."""
    from sbpeye.laws_links import find_law_references
    from test_laws_search import add_law, make_session

    db, _ = make_session()
    add_law(db, "sbp-act", title="State Bank of Pakistan Act, 1956",
            file_type="html", text_body="An Act to constitute a State Bank.", index=False)
    add_law(db, "lolr", title="Regulations for Lender of Last Resort Facility",
            file_type="html", index=False,
            text_body=(
                "In exercise of the powers conferred by section 17G of the State Bank of "
                "Pakistan Act, 1956, the State Bank is pleased to make these regulations."
            ))

    source = db.query(RegDocumentForTest).filter(RegDocumentForTest.id == "lolr").one()
    references = find_law_references(db, source)

    assert [ref.target_id for ref in references] == ["sbp-act"]
    # The evidence travels with the candidate, or the classifier has nothing to judge.
    assert "powers conferred by section 17G" in references[0].snippets[0]


def test_a_document_does_not_reference_itself_its_container_or_its_parts():
    from sbpeye.laws_links import find_law_references
    from test_laws_search import add_law, make_session

    db, _ = make_session()
    add_law(db, "fe-manual", title="Foreign Exchange Manual", file_type="manifest",
            text_body=None, index=False)
    add_law(db, "ch-12", title="EXPORTS", parent_id="fe-manual", part_label="Chapter 12",
            file_type="html", index=False,
            text_body="Under the Foreign Exchange Manual, exporters shall repatriate proceeds.")

    child = db.query(RegDocumentForTest).filter(RegDocumentForTest.id == "ch-12").one()

    # The container relationship is already modelled by parent_id; repeating it as an
    # edge would be noise on every chapter of every manual.
    assert find_law_references(db, child) == []


def test_relationship_generation_types_edges_and_keeps_unclassified_ones():
    from sqlalchemy.orm import sessionmaker

    import sbpeye.laws_ai as laws_ai_module
    from sbpeye.models import AIGenerationJob, RegDocumentRelationship
    from test_laws_search import add_law, make_session

    db, engine = make_session()
    add_law(db, "sbp-act", title="State Bank of Pakistan Act, 1956", file_type="html",
            text_body="An Act to constitute a State Bank.", index=False)
    add_law(db, "bco", title="Banking Companies Ordinance 1962", file_type="html",
            text_body="An Ordinance to regulate banking companies.", index=False)
    add_law(db, "lolr", title="Regulations for Lender of Last Resort Facility",
            file_type="html", index=False,
            text_body=(
                "In exercise of the powers conferred by section 17G of the State Bank of "
                "Pakistan Act, 1956, and having regard to a bank as defined in the "
                "Banking Companies Ordinance 1962, these regulations are made."
            ))
    db.add(AIGenerationJob(
        id="job-1", target_kind="law", document_id="lolr", feature="relationships",
        status="queued", created_at=datetime(2026, 8, 14),
    ))
    db.commit()

    client = FakeLawClient()
    # The model classifies one candidate and stays silent on the other.
    client.classify_law_references = lambda title, candidates: {
        "sbp-act": {"type": "made_under", "confidence": 0.9}
    }
    client.classify_circular_law_actions = lambda title, circulars: {}
    monkeypatch_target = sessionmaker(bind=engine, autoflush=False)
    laws_ai_module.SessionLocal = monkeypatch_target
    original_client = laws_ai_module.get_ai_client
    laws_ai_module.get_ai_client = lambda db: client
    try:
        laws_ai_module.run_law_generation_job("job-1")
    finally:
        laws_ai_module.get_ai_client = original_client

    db.expire_all()
    edges = {
        e.target_document_id: e
        for e in db.query(RegDocumentRelationship).filter(
            RegDocumentRelationship.source_document_id == "lolr"
        )
    }
    assert edges["sbp-act"].type == "made_under"
    assert edges["sbp-act"].detected_via == "ai"
    # A candidate the model skipped is still a real mention; it keeps the weaker,
    # deterministic reading rather than vanishing.
    assert edges["bco"].type == "references"
    assert edges["bco"].detected_via == "name_match"


def test_only_name_matched_circular_links_are_offered_for_typing():
    """A hyperlink into the document is stronger evidence than a model's opinion."""
    from sbpeye.laws_ai import _circular_link_candidates
    from sbpeye.models import Circular, RegDocumentLink
    from test_laws_search import add_law, make_session

    db, _ = make_session()
    document, _ = add_law(db, "doc-1", title="Prudential Regulations for SME Financing",
                          file_type="html", index=False)
    db.add(Circular(id="c-1", title="Amendments", reference="BPRD 1",
                    content_text="The Prudential Regulations for SME Financing are amended.",
                    date=datetime(2025, 1, 1)))
    db.add(Circular(id="c-2", title="Linked", reference="BPRD 2",
                    content_text="See the attached file.", date=datetime(2025, 2, 1)))
    db.add(RegDocumentLink(circular_id="c-1", document_id="doc-1",
                           link_type="references", detected_via="name_match"))
    db.add(RegDocumentLink(circular_id="c-2", document_id="doc-1",
                           link_type="references", detected_via="url_scan"))
    db.commit()
    db.refresh(document)

    candidates = _circular_link_candidates(document)

    assert [c["id"] for c in candidates] == ["c-1"]
    assert "are amended" in candidates[0]["snippets"][0]


def test_law_detail_carries_relationships_both_ways(law_client):
    from sbpeye.models import RegDocumentRelationship
    from test_laws_search import add_law

    test_client, db, _ = law_client
    add_law(db, "sbp-act", title="State Bank of Pakistan Act, 1956", file_type="html",
            index=False)
    add_law(db, "lolr", title="LOLR Regulations", file_type="html", index=False)
    db.add(RegDocumentRelationship(
        source_document_id="lolr", target_document_id="sbp-act",
        type="made_under", detected_via="ai", confidence=0.9,
        created_at=datetime(2026, 8, 14),
    ))
    db.commit()

    subordinate = test_client.get("/api/laws/lolr").json()
    parent = test_client.get("/api/laws/sbp-act").json()

    assert subordinate["relationships"]["outgoing"][0]["type"] == "made_under"
    assert subordinate["relationships"]["outgoing"][0]["document"]["id"] == "sbp-act"
    # And the Act knows what was made under it.
    assert parent["relationships"]["incoming"][0]["document"]["id"] == "lolr"


def test_the_law_relation_vocabulary_has_no_supersedes():
    """Between two laws, supersession is an edition change — versions already model it."""
    assert "supersedes" not in AIClient.LAW_RELATION_TYPES
    assert set(AIClient.LAW_RELATION_TYPES) == {
        "made_under", "amends", "repeals", "references"
    }


def test_an_invented_id_or_type_is_dropped_rather_than_written():
    client = AIClient(AIConfig())
    payload = json.dumps({"relations": [
        {"id": "real", "type": "made_under", "confidence": 0.9},
        {"id": "hallucinated", "type": "made_under", "confidence": 1.0},
        {"id": "real2", "type": "invented_type", "confidence": 1.0},
    ]})

    classified = client._parse_relations(
        payload, "relations", {"real", "real2"}, AIClient.LAW_RELATION_TYPES
    )

    assert list(classified) == ["real"]


def test_candidates_without_quotable_evidence_are_not_sent_to_the_model():
    """110 of 794 live edges match a canonical name that is not in the text as written."""
    from sbpeye.laws_ai import _circular_link_candidates
    from sbpeye.models import Circular, RegDocumentLink
    from test_laws_search import add_law, make_session

    db, _ = make_session()
    document, _ = add_law(db, "doc-1", title="Prudential Regulations for SME Financing",
                          file_type="html", index=False)
    db.add(Circular(id="c-1", title="Unrelated", reference="BPRD 9",
                    content_text="This circular has nothing quotable in it.",
                    date=datetime(2025, 1, 1)))
    db.add(RegDocumentLink(circular_id="c-1", document_id="doc-1",
                           link_type="references", detected_via="name_match"))
    db.commit()
    db.refresh(document)

    assert _circular_link_candidates(document) == []


def test_many_candidates_are_packed_into_several_calls():
    """One live document is named by 314 circulars; that is not one prompt."""
    from sbpeye.laws_ai import _pack_candidates

    candidates = [
        {"id": f"c-{i}", "label": "BPRD Circular", "snippets": ["x" * 500]}
        for i in range(20)
    ]
    batches = _pack_candidates(candidates, budget_chars=2_000)

    assert len(batches) > 1
    assert sum(len(batch) for batch in batches) == 20
    # Order is preserved, so a partial run is a prefix rather than a scatter.
    assert [c["id"] for batch in batches for c in batch] == [c["id"] for c in candidates]


def test_an_oversized_candidate_rides_alone_rather_than_being_dropped():
    from sbpeye.laws_ai import _pack_candidates

    big = {"id": "big", "label": "L", "snippets": ["x" * 5_000]}
    small = {"id": "small", "label": "L", "snippets": ["x" * 10]}

    batches = _pack_candidates([big, small], budget_chars=1_000)

    assert [[c["id"] for c in batch] for batch in batches] == [["big"], ["small"]]


def test_internal_page_markers_never_reach_a_prompt():
    """`[[SBPEYE_PAGE:n]]` is our own extraction bookkeeping, not the instrument's wording."""
    from sbpeye.laws_links import mention_snippets

    # As extraction writes them: on their own line.
    text = (
        "[[SBPEYE_PAGE:1]]\nRegulations made under Section 17G of the\n"
        "[[SBPEYE_PAGE:2]]\nState Bank of Pakistan Act, 1956 apply."
    )

    snippets = mention_snippets(text, "state bank of pakistan act 1956")

    assert snippets
    assert "SBPEYE_PAGE" not in snippets[0]
    assert "made under Section 17G" in snippets[0]


def test_a_snippet_window_that_slices_a_page_marker_still_cleans_it():
    """The window is cut at a character offset, so it can land mid-marker."""
    from sbpeye.laws_links import mention_snippets

    text = "[[SBPEYE_PAGE:12]]\n" + "x" * 400 + "\nthe Credit Bureau Act 2015 applies."

    snippets = mention_snippets(text, "credit bureau act 2015", window=420)

    assert "SBPEYE_PAGE" not in snippets[0]
    assert "PAGE:" not in snippets[0]


# --- phase F: container rollups and the CLI -----------------------------------------


def make_container(db, summarised_parts=2, unsummarised_parts=0):
    """A container whose parts carry their own version-level summaries."""
    from test_laws_search import add_law

    add_law(db, "fe-manual", title="Foreign Exchange Manual", file_type="manifest",
            text_body=None, index=False)
    for index in range(summarised_parts + unsummarised_parts):
        _, version = add_law(
            db, f"ch-{index}", title=f"CHAPTER {index}", parent_id="fe-manual",
            part_label=f"Chapter {index}", part_order=index,
            file_type="html", text_body="Rules.", index=False,
        )
        if index < summarised_parts:
            version.summary = f"Chapter {index} governs a thing."
    db.commit()
    from sbpeye.models import RegDocument
    return db.query(RegDocument).filter(RegDocument.id == "fe-manual").one()


def test_a_rollup_reads_parts_in_chapter_order_and_skips_unanalysed_ones():
    from sbpeye.laws_ai import rollup_sources
    from test_laws_search import make_session

    db, _ = make_session()
    container = make_container(db, summarised_parts=2, unsummarised_parts=1)

    sources = rollup_sources(container)

    # A manual read alphabetically is not a manual, and a rollup over the analysed third
    # of it would read as a summary of the whole.
    assert [label for label, _ in sources] == ["Chapter 0", "Chapter 1"]


def test_a_containers_rollup_is_stored_on_the_document_not_the_manifest():
    """The manifest's hash changes whenever any part does; the rollup must outlive that."""
    from sqlalchemy.orm import sessionmaker

    import sbpeye.laws_ai as laws_ai_module
    from sbpeye.models import AIGenerationJob, RegDocument
    from test_laws_search import make_session

    db, engine = make_session()
    make_container(db)
    db.add(AIGenerationJob(
        id="job-1", target_kind="law", document_id="fe-manual", feature="summary",
        status="queued", created_at=datetime(2026, 8, 14),
    ))
    db.commit()

    client = FakeLawClient()
    laws_ai_module.SessionLocal = sessionmaker(bind=engine, autoflush=False)
    original = laws_ai_module.get_ai_client
    laws_ai_module.get_ai_client = lambda db: client
    try:
        laws_ai_module.run_law_generation_job("job-1")
    finally:
        laws_ai_module.get_ai_client = original

    db.expire_all()
    container = db.query(RegDocument).filter(RegDocument.id == "fe-manual").one()
    assert container.summary == "Reduced summary of Foreign Exchange Manual."
    assert container.summary_generated_at is not None
    assert container.current_version.summary is None
    # The rollup was built from the parts, not from thin air.
    assert client.reduce_calls[0] == [
        "Chapter 0: Chapter 0 governs a thing.",
        "Chapter 1: Chapter 1 governs a thing.",
    ]


def test_the_api_reports_a_containers_rollup_as_its_summary(law_client):
    test_client, db, _ = law_client
    container = make_container(db)
    container.summary = "The manual governs foreign exchange."
    container.summary_generated_at = datetime(2026, 8, 14)
    db.commit()

    payload = test_client.get("/api/laws/fe-manual").json()

    # A container has a current version (the manifest), so a fallback would never fire —
    # the serializer has to select on shape.
    assert payload["summary"] == "The manual governs foreign exchange."
    assert payload["generation"]["summary"] is not None


def test_a_container_is_offered_the_features_it_can_actually_produce(law_client):
    test_client, db, started = law_client
    make_container(db)

    allowed = test_client.post("/api/laws/fe-manual/generate", json={"feature": "summary"})
    refused = test_client.post("/api/laws/fe-manual/generate", json={"feature": "checklist"})

    assert allowed.status_code == 202
    assert started == [allowed.json()["id"]]
    assert refused.status_code == 422
    assert refused.json()["structural"] is True
    assert "parts" in refused.json()["error"]


def test_a_container_with_no_analysed_parts_says_what_to_do_first(law_client):
    test_client, db, _ = law_client
    make_container(db, summarised_parts=0, unsummarised_parts=2)

    response = test_client.post("/api/laws/fe-manual/generate", json={"feature": "summary"})

    assert response.status_code == 422
    # Recoverable: summarise the parts and this becomes possible, so the UI should not
    # present it as a permanent fact about the document.
    assert response.json()["structural"] is False
    assert "parts first" in response.json()["error"]


def test_the_cli_exposes_one_command_per_feature():
    from click.testing import CliRunner

    from sbpeye.cli.commands import cli
    from sbpeye.laws_ai import LAW_GENERATION_FEATURES

    output = CliRunner().invoke(cli, ["laws", "--help"]).output

    for feature in LAW_GENERATION_FEATURES:
        name = "summarize" if feature == "summary" else feature
        assert name in output, f"`sbpeye laws {name}` is not registered"


def test_the_cli_skips_already_generated_documents_unless_forced(monkeypatch):
    from click.testing import CliRunner
    from sqlalchemy.orm import sessionmaker

    import sbpeye.cli.commands as cli_module
    from sbpeye.cli.commands import cli
    from sbpeye.models import RegDocument
    from test_laws_search import add_law, make_session

    db, engine = make_session()
    _, done = add_law(db, "done", title="Already analysed", file_type="html", index=False)
    done.summary = "Existing."
    done.summary_generated_at = datetime(2026, 8, 1)
    add_law(db, "todo", title="Not yet analysed", file_type="html", index=False)
    db.commit()

    client = FakeLawClient()
    monkeypatch.setattr(cli_module, "SessionLocal", sessionmaker(bind=engine, autoflush=False))
    monkeypatch.setattr(cli_module, "get_ai_client", lambda db: client)

    result = CliRunner().invoke(cli, ["laws", "summarize", "--delay", "0"])

    assert result.exit_code == 0, result.output
    assert "1 generated, 1 skipped" in result.output
    assert "already generated" in result.output
    db.expire_all()
    assert db.query(RegDocument).filter(
        RegDocument.id == "todo"
    ).one().current_version.summary is not None


def test_the_cli_reports_why_each_document_was_skipped(monkeypatch):
    """A batch run that silently does nothing is indistinguishable from a broken one."""
    from click.testing import CliRunner
    from sqlalchemy.orm import sessionmaker

    import sbpeye.cli.commands as cli_module
    from sbpeye.cli.commands import cli
    from test_laws_search import add_law, make_session

    db, engine = make_session()
    add_law(db, "ext", title="External law", file_type="html", is_external=1, index=False)
    add_law(db, "xls", title="Questionnaire", file_type="xls", text_body=None, index=False)
    db.commit()

    monkeypatch.setattr(cli_module, "SessionLocal", sessionmaker(bind=engine, autoflush=False))
    monkeypatch.setattr(cli_module, "get_ai_client", lambda db: FakeLawClient())

    result = CliRunner().invoke(cli, ["laws", "summarize", "--delay", "0"])

    assert "0 generated, 2 skipped" in result.output
    assert "published outside SBP" in result.output
    assert "spreadsheet" in result.output


def test_an_unknown_feature_raises_rather_than_quietly_producing_nothing():
    """The bug this guards: `summarize` (command) is not `summary` (feature)."""
    from sbpeye.laws_ai import LAW_GENERATION_FEATURES, _requested_features

    assert _requested_features("summary", LAW_GENERATION_FEATURES) == ("summary",)
    assert _requested_features("all", LAW_GENERATION_FEATURES) == LAW_GENERATION_FEATURES
    with pytest.raises(ValueError, match="Unknown feature"):
        _requested_features("summarize", LAW_GENERATION_FEATURES)


def test_generate_all_leaves_the_checklist_alone(monkeypatch):
    """The checklist is a call per chunk over a document that can run to hundreds of
    pages, so `all` skips it and it is asked for by name."""
    from sqlalchemy.orm import sessionmaker

    import sbpeye.laws_ai as laws_ai_module
    from sbpeye.models import AIGenerationJob, RegDocument
    from test_laws_search import add_law, make_session

    assert "checklist" not in laws_ai_module.LAW_BULK_FEATURES
    assert "checklist" in laws_ai_module.LAW_GENERATION_ACTIONS

    db, engine = make_session()
    add_law(db, "doc-1", file_type="html", index=False)
    db.add(AIGenerationJob(
        id="job-1", target_kind="law", document_id="doc-1", feature="all",
        status="queued", created_at=datetime(2026, 8, 14),
    ))
    db.commit()

    client = FakeLawClient()
    monkeypatch.setattr(laws_ai_module, "SessionLocal", sessionmaker(bind=engine, autoflush=False))
    monkeypatch.setattr(laws_ai_module, "get_ai_client", lambda db: client)

    laws_ai_module.run_law_generation_job("job-1")

    db.expire_all()
    stored = db.query(RegDocument).filter(RegDocument.id == "doc-1").one()
    assert stored.current_version.summary  # the rest of `all` still ran
    assert stored.current_version.compliance_checklist is None
    assert stored.current_version.checklist_generated_at is None

    # And asking for one by name still produces it.
    db.add(AIGenerationJob(
        id="job-2", target_kind="law", document_id="doc-1", feature="checklist",
        status="queued", created_at=datetime(2026, 8, 14),
    ))
    db.commit()
    laws_ai_module.run_law_generation_job("job-2")

    db.expire_all()
    stored = db.query(RegDocument).filter(RegDocument.id == "doc-1").one()
    assert json.loads(stored.current_version.compliance_checklist)["status"] == "completed"


def test_the_cli_summarises_parts_before_rolling_up_their_container(monkeypatch):
    """A container that sorts before its own parts must still roll up in one pass."""
    from click.testing import CliRunner
    from sqlalchemy.orm import sessionmaker

    import sbpeye.cli.commands as cli_module
    from sbpeye.cli.commands import cli
    from sbpeye.models import RegDocument
    from test_laws_search import add_law, make_session

    db, engine = make_session()
    # "AAA…" sorts before "ZZZ…", exactly as the Draft White Label ATM Guidelines sort
    # before their only chapter in the live corpus.
    add_law(db, "container", title="AAA Collection", file_type="manifest",
            text_body=None, index=False)
    add_law(db, "part", title="ZZZ Chapter", parent_id="container",
            part_label="Chapter 1", part_order=1, file_type="html",
            text_body="Rules.", index=False)
    db.commit()

    monkeypatch.setattr(cli_module, "SessionLocal", sessionmaker(bind=engine, autoflush=False))
    monkeypatch.setattr(cli_module, "get_ai_client", lambda db: FakeLawClient())

    result = CliRunner().invoke(cli, ["laws", "summarize", "--delay", "0"])

    assert "2 generated, 0 skipped" in result.output, result.output
    db.expire_all()
    container = db.query(RegDocument).filter(RegDocument.id == "container").one()
    assert container.summary is not None


# --- phase G: Excel export ----------------------------------------------------------


LAW_CHECKLIST = {
    "status": "completed",
    "generated_at": "2026-08-14T10:00:00",
    "coverage_gaps": [],
    "analysis_blocks": [{"block_id": "b1"}],
    "checklist_items": [{
        "doc_label": "sme-prs.pdf", "doc_type": "law", "ref": "Regulation SME R-1",
        "page_start": 4, "page_end": 4, "classification": "required",
        "requirement": "Banks shall obtain audited accounts.", "actor": "Banks",
        "applicability": "SME financing", "deadline": "", "evidence": "Audited accounts",
        "conditions": "", "source_text": "The bank shall obtain…",
    }],
    "source_units": [],
}


def test_a_law_exports_its_own_identity_columns_not_a_circulars(law_client):
    from openpyxl import load_workbook
    from io import BytesIO
    from test_laws_search import add_law

    test_client, db, _ = law_client
    _, version = add_law(db, "doc-1", title="Prudential Regulations for SME Financing (Updated till June 2024)",
                         file_type="pdf", index=False)
    version.version_label = "Updated till June 2024"
    version.compliance_checklist = json.dumps(LAW_CHECKLIST)
    db.commit()

    response = test_client.get("/api/laws/doc-1/checklist.xlsx")

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    headers = [cell.value for cell in workbook["Checklist"][1]]
    # A regulation is identified by its edition, not by a reference and a department.
    assert headers[:4] == ["Document Title", "Type", "Edition", "Captured"]
    assert "Circular Reference" not in headers
    row = [cell.value for cell in workbook["Checklist"][2]]
    assert row[0] == "Prudential Regulations for SME Financing"
    assert row[2] == "Updated till June 2024"
    assert row[9] == "required"
    assert row[10] == "Banks shall obtain audited accounts."
    # The edition is on the summary sheet too, with the hash that identifies it.
    fields = {r[0].value: r[1].value for r in workbook["Summary"].iter_rows(min_row=2)}
    assert fields["Edition"] == "Updated till June 2024"
    assert fields["Content Hash"] == "hash-doc-1"
    assert 'filename="Prudential_Regulations_for_SME_Financing_checklist.xlsx"' in \
        response.headers["content-disposition"]


def test_a_part_export_names_its_collection(law_client):
    from openpyxl import load_workbook
    from io import BytesIO
    from test_laws_search import add_law

    test_client, db, _ = law_client
    add_law(db, "fe-manual", title="Foreign Exchange Manual", file_type="manifest",
            text_body=None, index=False)
    _, version = add_law(db, "ch-12", title="EXPORTS", parent_id="fe-manual",
                         part_label="Chapter 12", part_order=12, file_type="pdf",
                         index=False)
    version.compliance_checklist = json.dumps(LAW_CHECKLIST)
    db.commit()

    response = test_client.get("/api/laws/ch-12/checklist.xlsx")

    workbook = load_workbook(BytesIO(response.content))
    fields = {r[0].value: r[1].value for r in workbook["Summary"].iter_rows(min_row=2)}
    # A part never appears without its container.
    assert fields["Part"] == "Chapter 12"
    assert fields["Collection"] == "Foreign Exchange Manual"


def test_a_checklist_from_a_superseded_edition_is_not_exported(law_client):
    from test_laws_search import add_law

    test_client, db, _ = law_client
    add_law(db, "doc-1", file_type="html", index=False)
    superseded = make_version("doc-1-old", "doc-1", is_current=0, file_type="html",
                              content_text="Old.", first_seen_at=datetime(2024, 1, 1),
                              compliance_checklist=json.dumps(LAW_CHECKLIST))
    db.add(superseded)
    db.commit()

    response = test_client.get("/api/laws/doc-1/checklist.xlsx")

    # It is on disk, but it is not what this document requires today.
    assert response.status_code == 404
    assert "does not have a generated checklist" in response.json()["error"]


def test_the_export_404s_for_an_unknown_document(law_client):
    test_client, _, _ = law_client
    assert test_client.get("/api/laws/nope/checklist.xlsx").status_code == 404


def test_a_circular_export_is_unchanged_by_the_generalisation():
    """The circular workbook is a shipped format; generalising must not move a column."""
    from io import BytesIO

    from openpyxl import load_workbook

    from sbpeye.checklist_export import build_checklist_workbook, circular_subject

    circular = SimpleNamespace(
        id="c-1", reference="BPRD/1", title="Reporting", department="BPRD",
        date=None, url="https://www.sbp.org.pk/rules.pdf",
    )

    workbook = load_workbook(build_checklist_workbook(circular, LAW_CHECKLIST))

    headers = [cell.value for cell in workbook["Checklist"][1]]
    assert headers[:4] == [
        "Circular Reference", "Circular Title", "Department", "Circular Date"
    ]
    assert workbook["Summary"]["B2"].value == "BPRD/1"
    assert circular_subject(circular).safe_filename() == "BPRD_1_checklist.xlsx"
